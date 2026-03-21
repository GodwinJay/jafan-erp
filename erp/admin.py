from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.utils import timezone
from django.utils.html import format_html, mark_safe
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect

# Import Unfold components
from unfold.admin import ModelAdmin, TabularInline
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    User, Material, BlockType, BusinessRules, PaymentAccount,
    Team, Machine, Customer, Site, Employee, Vendor, Truck,
    ExpenseCategory, Expense, ProcurementLog, ProductionLog,
    Payment, SalesOrder, SalesOrderItem, SupplyLog,
    ReturnLog, CashRefund, BreakageLog, FuelLog, MaintenanceLog,
    TransportAsset, TransportRevenue, BankCharge, AccountTransfer,
    DailyCashClose, VendorPayment, TeamPayment, SandVehicleType, SandSale, Debtor, Loan, LoanRepayment, QuickSale
)


# =============================================================================
# AJAX VIEWS FOR CHAINED DROPDOWNS
# =============================================================================

def get_customer_sites(request):
    """Get sites for a specific customer."""
    customer_id = request.GET.get('customer_id')
    if customer_id:
        sites = Site.objects.filter(customer_id=customer_id).values('id', 'name')
        return JsonResponse(list(sites), safe=False)
    return JsonResponse([], safe=False)


def get_customer_orders(request):
    """Get pending/partial sales orders for a specific customer."""
    customer_id = request.GET.get('customer_id')
    if customer_id:
        orders = SalesOrder.objects.filter(
            customer_id=customer_id,
            status__in=['PENDING', 'PARTIAL']
        ).values('id', 'date', 'pk')
        result = [{'id': o['id'], 'name': f"SO-{o['pk']:05d} ({o['date']})"} for o in orders]
        return JsonResponse(result, safe=False)
    return JsonResponse([], safe=False)


def get_order_items(request):
    """Get unfulfilled items for a specific sales order."""
    order_id = request.GET.get('order_id')
    if order_id:
        items = SalesOrderItem.objects.filter(order_id=order_id).select_related('block_type')
        result = []
        for item in items:
            remaining = item.quantity_requested - item.quantity_supplied
            if remaining > 0:
                result.append({
                    'id': item.id,
                    'name': f"{item.block_type.name} - {remaining} remaining @ ₦{item.agreed_price}"
                })
        return JsonResponse(result, safe=False)
    return JsonResponse([], safe=False)


def get_vendor_materials(request):
    """Get materials supplied by a specific vendor."""
    vendor_id = request.GET.get('vendor_id')
    if vendor_id:
        vendor = Vendor.objects.filter(pk=vendor_id).first()
        if vendor and vendor.supply_type:
            material_map = {
                'CEMENT': ['CEMENT'],
                'SAND': ['SHARP_SAND', 'BLACK_SAND'],
                'TRANSPORT': [],
            }
            material_names = material_map.get(vendor.supply_type, [])
            materials = Material.objects.filter(name__in=material_names).values('id', 'name')
            result = [{'id': m['id'], 'name': m['name']} for m in materials]
            return JsonResponse(result, safe=False)
    materials = Material.objects.all().values('id', 'name')
    return JsonResponse(list(materials), safe=False)


# =============================================================================
# RESTRICTED ADMIN MIXIN
# =============================================================================

class RestrictedAdmin(ModelAdmin):
    """
    Base admin class with role-based permissions.
    
    Roles:
    - ADMIN (Jeremiah): Full access - view, add, edit, delete everything
    - MANAGER (Ezekiel): View & add everything, no edit/delete, can approve loans
    - OPERATIONS (Joshua): Production, Inventory, Fuel, Sand Sales, Loans
    - SITE_MANAGER (Ene): Sales, Customers, Expenses, Cash, Reports
    - SALES (Nkiruka): Sales Orders, Payments, Customers (limited)
    - TRANSPORT (Daniel): Transport, Fuel, Maintenance, Vendor Payments
    
    Note: Everyone can do sales (Sales Orders, Payments, Customers, Sites)
    """
    
    # Common sales models everyone can access
    COMMON_SALES_MODELS = [
        'customer', 'site', 'salesorder', 'salesorderitem', 
        'supplylog', 'payment', 'blocktype', 'quicksale'
    ]
    
    # Define which models each role can access (in addition to common sales)
    ROLE_MODEL_ACCESS = {
        'ADMIN': '__all__',
        
        'MANAGER': '__all__',  # Can view/add everything
        
        'OPERATIONS': [
            # Production & Inventory
            'productionlog', 'material', 'procurementlog', 'machine',
            'team', 'employee', 'teampayment',
            # Stock & Breakages
            'returnlog', 'cashrefund',
            # Fuel & Diesel
            'fuellog', 'truck',
            # Sand Sales
            'sandsale', 'sandvehicletype',
            # Loans (sand-related)
            'loan', 'loanrepayment', 'debtor',
            # Transport Dashboard (view trips)
            'transportrevenue',
        ],
        
        'SITE_MANAGER': [
            # Sales (already in common)
            'cashrefund', 'returnlog',
            # Expenses & Cash
            'expense', 'expensecategory', 'paymentaccount',
            # Monitoring
            'fuellog',
            # Vendors
            'vendor', 'vendorpayment',
            # Reports access (dashboards handled separately)
        ],
        
        'SALES': [
            # Basic sales (already in common)
            'cashrefund', 'returnlog',
            # Expenses & Cash
            'expense', 'expensecategory', 'paymentaccount',
        ],
        
        'TRANSPORT': [
            # Transport operations
            'transportrevenue', 'fuellog', 'maintenancelog',
            'truck', 'transportasset',
            # Drivers
            'employee',
            # Vendor payments (spare parts)
            'vendor', 'vendorpayment',
        ],
    }
    
    # Roles that can EDIT records (not just view/add)
    EDIT_ROLES = ['ADMIN']
    
    # Roles that can DELETE records
    DELETE_ROLES = ['ADMIN']
    
    # Roles that can APPROVE loans
    LOAN_APPROVAL_ROLES = ['ADMIN', 'MANAGER']

    def _get_model_name(self):
        """Get the lowercase model name for this admin."""
        return self.model._meta.model_name

    def _get_user_role(self, user):
        """Get the role of the user from their employee profile."""
        if user.is_superuser:
            return 'ADMIN'
        try:
            if hasattr(user, 'employee_profile') and user.employee_profile:
                return user.employee_profile.role or ''
        except:
            pass
        return ''

    def _user_has_role_access(self, user):
        """Check if user's role grants access to this model."""
        if user.is_superuser:
            return True
        
        role = self._get_user_role(user)
        if not role:
            return False
        
        # ADMIN and MANAGER have access to everything
        if role in ['ADMIN', 'MANAGER']:
            return True
        
        model_name = self._get_model_name()
        
        # Everyone can access common sales models
        if model_name in self.COMMON_SALES_MODELS:
            return True
        
        # Check role-specific access
        allowed_models = self.ROLE_MODEL_ACCESS.get(role, [])
        return model_name in allowed_models

    def has_module_permission(self, request):
        """Determines if the user can see the app in admin."""
        return self._user_has_role_access(request.user)

    def has_view_permission(self, request, obj=None):
        return self._user_has_role_access(request.user)

    def has_add_permission(self, request):
        return self._user_has_role_access(request.user)

    def has_change_permission(self, request, obj=None):
        """Only ADMIN can edit records."""
        role = self._get_user_role(request.user)
        return request.user.is_superuser or role in self.EDIT_ROLES

    def has_delete_permission(self, request, obj=None):
        """Only ADMIN can delete records."""
        role = self._get_user_role(request.user)
        return request.user.is_superuser or role in self.DELETE_ROLES

    def delete_model(self, request, obj):
        obj.delete()

    def delete_queryset(self, request, queryset):
        for obj in queryset:
            obj.delete()
    
    def can_approve_loan(self, user):
        """Check if user can approve loans."""
        role = self._get_user_role(user)
        return role in self.LOAN_APPROVAL_ROLES


# =============================================================================
# BULK EXPORT ACTIONS
# =============================================================================

def export_to_excel(modeladmin, request, queryset):
    """Generic export action for any model."""
    model = queryset.model
    model_name = model._meta.verbose_name_plural.title().replace(" ", "_")

    wb = Workbook()
    ws = wb.active
    ws.title = model_name[:30]

    exclude_fields = ['id', 'created_at', 'updated_at']
    fields = [f for f in model._meta.fields if f.name not in exclude_fields]

    headers = [f.verbose_name.title() for f in fields]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field.name)
            if hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool, type(None))):
                value = str(value)
            display_method = f'get_{field.name}_display'
            if hasattr(obj, display_method) and callable(getattr(obj, display_method)):
                try: value = getattr(obj, display_method)()
                except: pass
            row.append(value)
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response

export_to_excel.short_description = "📊 Export selected to Excel"


def export_supplies_detailed(modeladmin, request, queryset):
    """Detailed export for Supply Logs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Supply_Log_Export"
    headers = ['Date', 'Customer', 'Block', 'Qty', 'Total']
    ws.append(headers)
    for log in queryset:
        ws.append([str(log.date), log.customer.name, log.block_type.name, log.quantity_delivered, log.total_value])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Supply_Logs.xlsx"'
    wb.save(response)
    return response
export_supplies_detailed.short_description = "📊 Export selected (Detailed) to Excel"

def export_expenses_detailed(modeladmin, request, queryset):
    return export_to_excel(modeladmin, request, queryset)
export_expenses_detailed.short_description = "📊 Export selected (Detailed) to Excel"

def export_customers_with_balance(modeladmin, request, queryset):
    return export_to_excel(modeladmin, request, queryset)
export_customers_with_balance.short_description = "📊 Export selected to Excel"

def export_production_detailed(modeladmin, request, queryset):
    return export_to_excel(modeladmin, request, queryset)
export_production_detailed.short_description = "📊 Export selected (Detailed) to Excel"

def export_payments_detailed(modeladmin, request, queryset):
    return export_to_excel(modeladmin, request, queryset)
export_payments_detailed.short_description = "📊 Export selected to Excel"

def export_procurement_detailed(modeladmin, request, queryset):
    return export_to_excel(modeladmin, request, queryset)
export_procurement_detailed.short_description = "📊 Export selected to Excel"

def export_returns_detailed(modeladmin, request, queryset):
    return export_to_excel(modeladmin, request, queryset)
export_returns_detailed.short_description = "📊 Export selected (Detailed) to Excel"


# =============================================================================
# ADMIN CLASSES
# =============================================================================

@admin.register(User)
class UserAdmin(BaseUserAdmin, ModelAdmin):
    fieldsets = BaseUserAdmin.fieldsets + (
        ("Role Information", {"fields": ("role",)}),
    )
    list_display = ["username", "email", "role", "is_active", "is_staff", "date_joined"]
    list_filter = ["role", "is_active"]


@admin.register(Material)
class MaterialAdmin(RestrictedAdmin):
    actions = [export_to_excel]
    list_display = ["name", "current_stock", "unit_price", "is_inventory_tracked", "is_low_stock"]
    list_filter = ["name", "is_inventory_tracked"]
    search_fields = ["name"] 
    
    def is_low_stock(self, obj):
        return obj.is_low_stock
    is_low_stock.boolean = True


@admin.register(BlockType)
class BlockTypeAdmin(RestrictedAdmin):
    actions = [export_to_excel]
    list_display = ["name", "current_stock", "weighted_average_cost", "selling_price", "is_active"]
    readonly_fields = ["weighted_average_cost"]
    search_fields = ["name"]
    
    fieldsets = (
        ("Inventory & Pricing", {
            "fields": ("name", "current_stock", "low_stock_threshold", "selling_price", "is_active")
        }),
        ("Costing (Auto-Calculated)", {
            "fields": ("weighted_average_cost",),
            "description": "This is the current value of your stock based on production costs."
        }),
        ("Production Recipe", {
            "fields": ("blocks_per_bag", "sand_ratio", "batch_size")
        }),
        ("Variable Costs (Per Block)", {
            "fields": ("operator_rate", "loader_rate", "stacking_rate", "logistics_rate")
        }),
    )


@admin.register(BusinessRules)
class BusinessRulesAdmin(RestrictedAdmin):
    list_display = ["name", "sand_cost", "diesel_power_cost", "updated_at"]
    
    def has_add_permission(self, request):
        return not BusinessRules.objects.exists()


@admin.action(description="🔍 Cash Audit - Record physical count")
def cash_audit_action(modeladmin, request, queryset):
    if queryset.count() != 1:
        messages.error(request, "Please select exactly ONE account.")
        return
    account = queryset.first()
    if account.account_type != 'CASH':
        messages.warning(request, f"Typically for CASH accounts. {account.bank_name} is {account.get_account_type_display()}.")
    return redirect(reverse('admin:erp_paymentaccount_change', args=[account.pk]))


@admin.register(PaymentAccount)
class PaymentAccountAdmin(RestrictedAdmin):
    list_display = [
        "bank_name", "account_type", "business_unit", "account_number", 
        "current_balance_display", "last_audit_date", "is_active", "pdf_actions"
    ]
    list_filter = ["account_type", "business_unit", "is_active"]
    search_fields = ["bank_name", "account_name", "account_number"]
    ordering = ["business_unit", "bank_name"]
    readonly_fields = ['current_balance', 'last_audit_balance', 'last_audit_notes']
    
    fieldsets = (
        (None, {
            'fields': ('bank_name', 'account_name', 'account_number', 'account_type', 'business_unit')
        }),
        ('Balance Tracking', {
            'fields': ('opening_balance', 'current_balance'),
            'description': 'Set opening balance ONCE when system goes live. Current balance updates automatically.'
        }),
        ('Daily Cash Audit', {
            'fields': ('last_audit_date', 'last_audit_balance', 'last_audit_notes'),
            'description': 'Record daily physical cash count.'
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )

    def has_module_permission(self, request):
        # Hide from sidebar for non-superusers
        return request.user.is_superuser

    def has_view_permission(self, request, obj=None):
        # Allow view for autocomplete to work
        return True

    def has_add_permission(self, request):
        return request.user.is_superuser

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser
    
    def get_queryset(self, request):
        # Allow all users to see active accounts (needed for autocomplete)
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        # Non-superusers only see active accounts
        return qs.filter(is_active=True)

    def formfield_for_choice_field(self, db_field, request, **kwargs):
        if db_field.name == "account_type" and not request.user.is_superuser:
            kwargs['choices'] = [(k, v) for k, v in PaymentAccount.ACCOUNT_TYPE_CHOICES if k != 'CASH']
        return super().formfield_for_choice_field(db_field, request, **kwargs)

    @admin.display(description="Current Balance")
    def current_balance_display(self, obj):
        amount = f"₦{obj.current_balance:,.2f}"
        if obj.current_balance >= 0:
            return format_html('<span style="color: green;">{}</span>', amount)
        return format_html('<span style="color: red;">{}</span>', amount)

    def pdf_actions(self, obj):
        statement_url = reverse('select_account_statement_date', args=[obj.pk])
        return format_html(
            '<a href="{}" style="padding:3px 8px; background:#28a745; color:white; border-radius:4px; text-decoration:none; font-size:11px;">📊 Statement</a>',
            statement_url
        )
    pdf_actions.short_description = "PDF"
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_statement_button'] = True
        if object_id:
            extra_context['statement_url'] = reverse('select_account_statement_date', args=[object_id])
        return super().change_view(request, object_id, form_url, extra_context)


@admin.register(DailyCashClose)
class DailyCashCloseAdmin(RestrictedAdmin):
    list_display = ['date', 'account', 'physical_cash_count', 'system_balance_display', 'difference_display', 'status_badge', 'closed_by']
    list_filter = ['date', 'status', 'account']
    readonly_fields = ['system_balance_at_close', 'difference', 'status', 'closed_by']

    def get_fieldsets(self, request, obj=None):
        if request.user.is_superuser:
            return (
                ('Closing Details', {'fields': ('date', 'account', 'closed_by')}),
                ('The Count', {'fields': ('physical_cash_count',)}),
                ('Audit Results', {'fields': ('system_balance_at_close', 'difference', 'status')}),
                ('Notes', {'fields': ('notes',)}),
            )
        return (
            ('Closing Details', {'fields': ('date', 'account')}),
            ('The Count', {'fields': ('physical_cash_count', 'notes')}),
        )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser: return qs
        return qs.none()

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.closed_by = request.user
            obj.system_balance_at_close = obj.account.current_balance
        super().save_model(request, obj, form, change)

    def response_add(self, request, obj, post_url_continue=None):
        if not request.user.is_superuser:
            messages.success(request, "✅ Cash count recorded successfully.")
            return redirect("admin:index")
        return super().response_add(request, obj, post_url_continue)

    @admin.display(description="System Says")
    def system_balance_display(self, obj):
        return f"₦{obj.system_balance_at_close:,.2f}"

    @admin.display(description="Discrepancy")
    def difference_display(self, obj):
        if obj.difference == 0:
            return format_html('<span style="color: #28a745; font-weight:bold;">{}</span>', "Balanced")
        elif obj.difference < 0:
            amount = f"{abs(obj.difference):,.2f}"
            return format_html('<span style="color: #dc3545; font-weight:bold;">-₦{} (Short)</span>', amount)
        else:
            amount = f"{obj.difference:,.2f}"
            return format_html('<span style="color: #d4af37; font-weight:bold;">+₦{} (Excess)</span>', amount)

    @admin.display(description="Status")
    def status_badge(self, obj):
        colors = {'BALANCED': 'green', 'SHORT': 'red', 'EXCESS': 'orange'}
        clean_label = obj.get_status_display().replace('✅ ', '').replace('⚠️ ', '').replace('❌ ', '')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; border-radius: 10px; font-size: 11px;">{}</span>',
            colors.get(obj.status, 'gray'), clean_label
        )


class SiteInline(TabularInline):
    model = Site
    extra = 1
    tab = True
    fields = ["name", "address", "contact_person", "contact_phone", "is_outside_town", "blocks_owed", "is_active"]


@admin.register(Customer)
class CustomerAdmin(RestrictedAdmin):
    actions = [export_customers_with_balance]
    list_display = ["name", "phone", "account_balance", "credit_limit", "balance_status", "total_blocks_owed", "is_active", "pdf_actions"]
    list_filter = ["customer_type", "is_active", "created_at"]
    search_fields = ["name", "phone", "email", "office_address"]
    ordering = ["name"]
    date_hierarchy = "created_at"
    inlines = [SiteInline]

    fieldsets = (
        ("Basic Info", {"fields": ("name", "phone", "email", "customer_type", "is_active")}),
        ("Address", {"fields": ("office_address",)}),
        ("Financial", {"fields": ("account_balance", "credit_limit")}),
        ("Notes", {"fields": ("notes",)}),
    )

    def balance_status(self, obj):
        return obj.balance_status
    balance_status.short_description = "Status"

    def total_blocks_owed(self, obj):
        return obj.total_blocks_owed
    total_blocks_owed.short_description = "Blocks Owed"

    def pdf_actions(self, obj):
        statement_url = reverse('select_statement_date', args=[obj.pk])
        return format_html(
            '<a href="{}" style="padding:3px 8px; background:#6A1B9A; color:white; border-radius:4px; text-decoration:none; font-size:11px;">📊 Statement</a>',
            statement_url
        )
    pdf_actions.short_description = "PDF"


@admin.register(ProductionLog)
class ProductionLogAdmin(RestrictedAdmin):
    actions = [export_production_detailed]
    list_display = ["date", "team", "machine", "block_type", "quantity_produced", "breakages", "team_pay", "unit_cost", "recorded_by"]
    list_filter = ["date", "team", "machine", "block_type"]
    search_fields = ["team__name", "machine__name", "block_type__name", "notes"]
    date_hierarchy = "date"
    readonly_fields = ["sharp_sand_used", "labor_cost", "unit_cost"]
    ordering = ["-date", "-created_at"]
    autocomplete_fields = ["team", "machine", "block_type"]

    fieldsets = (
        ("Production Details", {"fields": ("date", "team", "machine", "block_type")}),
        ("Output", {"fields": ("quantity_produced", "breakages")}),
        ("Materials Used (Manual)", {"fields": ("cement_used", "black_sand_used")}),
        ("Financials (Auto)", {"fields": ("sharp_sand_used", "labor_cost", "unit_cost")}),
        ("Notes", {"fields": ("notes",)}),
    )

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


class SalesOrderItemInline(TabularInline):
    model = SalesOrderItem
    extra = 1
    tab = True 
    fields = ["block_type", "quantity_requested", "discount_per_block", "discount_reason", "agreed_price", "quantity_supplied", "line_total_display"]
    readonly_fields = ["agreed_price", "quantity_supplied", "line_total_display"]
    autocomplete_fields = ["block_type"]

    def line_total_display(self, obj):
        if obj.pk: return f"₦{obj.line_total:,.2f}"
        return "-"
    line_total_display.short_description = "Line Total"


@admin.register(SalesOrder)
class SalesOrderAdmin(RestrictedAdmin):
    list_display = [
        "__str__", "date", "customer", "status",
        "supply_progress_display", "total_value_display",
        "valid_until", "pdf_actions"
    ]
    list_filter = ["status", "date", "customer"]
    search_fields = ["customer__name", "customer__phone", "site__name"]
    date_hierarchy = "date"
    ordering = ["pk"]
    inlines = [SalesOrderItemInline]
    autocomplete_fields = ["customer"]
    readonly_fields = ["valid_until"]

    fieldsets = (
        ("Order Info", {"fields": ("date", "customer", "site", "status")}),
        ("Pricing Adjustments", {
            "fields": ("surcharge_per_block",),
            "description": "Surcharge added PER BLOCK (for logistics). Discounts are set per item below."
        }),
        ("Validity", {"fields": ("valid_until",)}),
        ("Notes", {"fields": ("notes",)}),
    )

    class Media:
        js = ('/static/admin/js/sales_order_dropdowns.js',)

    def get_form(self, request, obj=None, **kwargs):
        self.obj_instance = obj
        return super().get_form(request, obj, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "site":
            if request.method == 'POST' and request.POST.get('customer'):
                kwargs["queryset"] = Site.objects.filter(customer_id=request.POST.get('customer'))
            elif hasattr(self, 'obj_instance') and self.obj_instance and self.obj_instance.customer_id:
                kwargs["queryset"] = Site.objects.filter(customer_id=self.obj_instance.customer_id)
            else:
                kwargs["queryset"] = Site.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def pdf_actions(self, obj):
        if obj.status == 'COMPLETED':
            url = reverse('generate_proforma', args=[obj.pk]) 
            label = "📄 Final Invoice"
            color = "#2E7D32"
        else:
            url = reverse('generate_proforma', args=[obj.pk])
            label = "📋 Proforma"
            color = "#D4AF37"

        return format_html(
            '<a href="{}" target="_blank" style="padding:3px 8px; background:{}; color:white; border-radius:4px; text-decoration:none; font-size:11px; font-weight:bold;">{}</a>',
            url, color, label
        )
    pdf_actions.short_description = "PDF"

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)
    
    def supply_progress_display(self, obj):
        p = obj.supply_progress
        color = "green" if p == 100 else "orange" if p > 0 else "gray"
        return format_html('<span style="color: {}; font-weight: bold;">{}%</span>', color, p)
    supply_progress_display.short_description = "Progress"

    def total_value_display(self, obj):
        return f"₦{obj.total_value:,.2f}"
    total_value_display.short_description = "Total Value"


@admin.register(SalesOrderItem)
class SalesOrderItemAdmin(RestrictedAdmin):
    list_display = ["order", "block_type", "quantity_requested", "agreed_price", "quantity_supplied"]
    list_filter = ["block_type", "order__customer"]
    search_fields = ["order__customer__name", "block_type__name"]
    autocomplete_fields = ["order", "block_type"]

    def has_module_permission(self, request):
        return False


@admin.register(Payment)
class PaymentAdmin(RestrictedAdmin):
    actions = [export_payments_detailed]
    list_display = ["date", "customer", "amount", "method", "sales_order", "recorded_by", "pdf_actions"]
    list_filter = ["date", "method", "payment_account", "customer"]
    search_fields = ["customer__name", "reference", "remark"]
    date_hierarchy = "date"
    ordering = ["-date", "-created_at"]
    autocomplete_fields = ["customer", "payment_account"]

    class Media:
        js = ('/static/admin/js/payment_dropdowns.js',)

    fieldsets = (
        ("Payment Info", {"fields": ("date", "customer", "amount")}),
        ("Link to Order (Optional)", {"fields": ("sales_order",)}),
        ("Details", {"fields": ("method", "payment_account", "reference", "remark")}),
    )

    def get_form(self, request, obj=None, **kwargs):
        self.obj_instance = obj
        return super().get_form(request, obj, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "sales_order":
            if request.method == 'POST' and request.POST.get('customer'):
                kwargs["queryset"] = SalesOrder.objects.filter(customer_id=request.POST.get('customer'))
            elif hasattr(self, 'obj_instance') and self.obj_instance and self.obj_instance.customer_id:
                kwargs["queryset"] = SalesOrder.objects.filter(customer_id=self.obj_instance.customer_id)
            else:
                kwargs["queryset"] = SalesOrder.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def pdf_actions(self, obj):
        receipt_url = reverse('generate_receipt', args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" style="padding:3px 8px; background:#2E7D32; color:white; border-radius:4px; text-decoration:none; font-size:11px;">🧾 Receipt</a>',
            receipt_url
        )
    pdf_actions.short_description = "PDF"

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(SupplyLog)
class SupplyLogAdmin(RestrictedAdmin):
    actions = [export_supplies_detailed]
    
    list_filter = ["date", "delivery_type", "customer", "site", "block_type", "truck"]
    search_fields = ["customer__name", "site__name", "remark"]
    date_hierarchy = "date"
    ordering = ["-date", "-created_at"]
    # Only autocomplete for non-chained fields
    autocomplete_fields = ["customer", "truck", "driver", "block_type"]
    readonly_fields = ["quantity_delivered", "total_value", "unit_price", "logistics_income", "cost_of_goods_sold", "gross_profit_on_sale"]

    class Media:
        js = ('/static/admin/js/chained_dropdowns.js',)

    def get_list_display(self, request):
        cols = [
            "date", "delivery_type", "customer", "site", "block_type",
            "quantity_loaded", "quantity_returned", "quantity_delivered",
            "total_value_display", "logistics_income_display"
        ]
        if request.user.is_superuser:
            cols.extend(["cogs_display", "gross_profit_display"])
        cols.extend(["truck", "pdf_actions"])
        return cols

    def get_fieldsets(self, request, obj=None):
        fieldsets = [
            ("Delivery Info", {"fields": ("date", "delivery_type", "customer", "site")}),
            ("Product & Order", {"fields": ("sales_order", "order_item", "block_type")}),
            ("Quantities", {
                "fields": ("quantity_loaded", "breakages", "quantity_returned", "quantity_delivered"),
                "description": "Delivered = Loaded - Breakages - Returned"
            }),
            ("Logistics", {"fields": ("truck", "driver", "pickup_authorized_by")}),
            ("Financials", {"fields": ("unit_price", "logistics_discount", "total_value", "logistics_income")}),
        ]
        if request.user.is_superuser:
            fieldsets.append(("Profitability (Auto-Calculated)", {
                "fields": ("cost_of_goods_sold", "gross_profit_on_sale"),
                "description": "Visible to Admin Only. COGS = Qty × WAC at time of sale."
            }))
        fieldsets.append(("Notes", {"fields": ("remark",)}))
        return fieldsets

    @admin.display(description="Total Value", ordering='total_value')
    def total_value_display(self, obj):
        formatted_amount = f"{obj.total_value:,.2f}"
        return f"₦{formatted_amount}"

    @admin.display(description="Gross Profit")
    def gross_profit_display(self, obj):
        formatted_amount = f"{obj.gross_profit_on_sale:,.2f}"
        if obj.gross_profit_on_sale >= 0:
            return format_html('<span style="color: green;">₦{}</span>', formatted_amount)
        return format_html('<span style="color: red;">₦{}</span>', formatted_amount)

    @admin.display(description="Logistics Income")
    def logistics_income_display(self, obj):
        if obj.logistics_income:
            return f"₦{obj.logistics_income:,.2f}"
        return "₦0.00"

    @admin.display(description="COGS")
    def cogs_display(self, obj):
        return f"₦{obj.cost_of_goods_sold:,.2f}"

    def pdf_actions(self, obj):
        invoice = reverse('generate_invoice', args=[obj.pk])
        waybill = reverse('generate_waybill', args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" style="margin-right:5px; color:green;">📄 Invoice</a>'
            '<a href="{}" target="_blank" style="color:blue;">🚚 Waybill</a>',
            invoice, waybill
        )
    pdf_actions.short_description = "PDFs"
    
    def save_model(self, request, obj, form, change):
        if not obj.recorded_by:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "site":
            # Check if we're editing an existing object or if customer is in POST data
            if request.method == 'POST' and request.POST.get('customer'):
                customer_id = request.POST.get('customer')
                kwargs["queryset"] = Site.objects.filter(customer_id=customer_id)
            elif hasattr(self, 'obj_instance') and self.obj_instance and self.obj_instance.customer_id:
                kwargs["queryset"] = Site.objects.filter(customer_id=self.obj_instance.customer_id)
            else:
                kwargs["queryset"] = Site.objects.none()
            
        elif db_field.name == "sales_order":
            if request.method == 'POST' and request.POST.get('customer'):
                customer_id = request.POST.get('customer')
                kwargs["queryset"] = SalesOrder.objects.filter(
                    customer_id=customer_id
                ).exclude(status__in=['COMPLETED', 'CANCELLED']).order_by('-date')
            elif hasattr(self, 'obj_instance') and self.obj_instance and self.obj_instance.customer_id:
                kwargs["queryset"] = SalesOrder.objects.filter(
                    customer_id=self.obj_instance.customer_id
                ).exclude(status__in=['COMPLETED', 'CANCELLED']).order_by('-date')
            else:
                kwargs["queryset"] = SalesOrder.objects.exclude(
                    status__in=['COMPLETED', 'CANCELLED']
                ).order_by('-date')
                
        elif db_field.name == "order_item":
            if request.method == 'POST' and request.POST.get('sales_order'):
                order_id = request.POST.get('sales_order')
                kwargs["queryset"] = SalesOrderItem.objects.filter(order_id=order_id)
            elif hasattr(self, 'obj_instance') and self.obj_instance and self.obj_instance.sales_order_id:
                kwargs["queryset"] = SalesOrderItem.objects.filter(order_id=self.obj_instance.sales_order_id)
            else:
                kwargs["queryset"] = SalesOrderItem.objects.none()
                
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_form(self, request, obj=None, **kwargs):
        # Store the object instance for use in formfield_for_foreignkey
        self.obj_instance = obj
        return super().get_form(request, obj, **kwargs)


@admin.register(Team)
class TeamAdmin(RestrictedAdmin):
    list_display = ["name", "is_active"]
    search_fields = ["name"] 


@admin.register(Machine)
class MachineAdmin(RestrictedAdmin):
    list_display = ["name", "machine_type", "status"]
    search_fields = ["name"] 


@admin.register(Site)
class SiteAdmin(RestrictedAdmin):
    list_display = ["customer", "name", "blocks_owed", "is_active"]
    search_fields = ["name", "address", "customer__name"] 
    autocomplete_fields = ["customer"]


@admin.register(Employee)
class EmployeeAdmin(RestrictedAdmin):
    list_display = ['name', 'phone', 'role', 'user', 'team', 'is_active']
    list_filter = ['role', 'is_active', 'team']
    search_fields = ['name', 'phone']
    autocomplete_fields = ['user', 'team']
    
    fieldsets = (
        ('Basic Info', {
            'fields': ('name', 'phone', 'role', 'user')
        }),
        ('Work', {
            'fields': ('team', 'pay_type', 'current_balance')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    ) 


@admin.register(Vendor)
class VendorAdmin(RestrictedAdmin):
    actions = [export_to_excel]
    list_display = [
        "name", "supply_type", "is_internal", "phone", 
        "account_balance_display", "balance_status", "is_active", "pdf_actions"
    ]
    list_filter = ["supply_type", "is_internal", "is_active"]
    search_fields = ["name", "phone", "address"]
    ordering = ["name"]
    readonly_fields = ['account_balance']

    fieldsets = (
        (None, {
            'fields': ('name', 'phone', 'address', 'supply_type')
        }),
        ('Internal Vendor', {
            'fields': ('is_internal',),
            'description': 'Check for internal departments like Jafan Transport'
        }),
        ('Credit Tracking', {
            'fields': ('account_balance',),
            'description': 'Positive = We owe them. Updates automatically from credit purchases.'
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )

    def account_balance_display(self, obj):
        if obj.account_balance > 0:
            amount = f"₦{obj.account_balance:,.2f}"
            return format_html('<span style="color: red;">{}</span>', amount)
        elif obj.account_balance < 0:
            amount = f"₦{abs(obj.account_balance):,.2f}"
            return format_html('<span style="color: blue;">{}</span>', amount)
        else:
            return format_html('<span style="color: green;">{}</span>', "₦0.00")
    
    @admin.display(description="Status")
    def balance_status(self, obj):
        if obj.account_balance > 0:
            return format_html('<span style="color: red;">{}</span>', "⚠️ We Owe")
        elif obj.account_balance < 0:
            return format_html('<span style="color: blue;">{}</span>', "They Owe")
        return format_html('<span style="color: green;">{}</span>', "✓ Settled")

    def pdf_actions(self, obj):
        statement_url = reverse('select_vendor_statement_date', args=[obj.pk])
        return format_html(
            '<a href="{}" style="padding:3px 8px; background:#6A1B9A; color:white; border-radius:4px; text-decoration:none; font-size:11px;">📊 Statement</a>',
            statement_url
        )
    pdf_actions.short_description = "PDF"

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_statement_button'] = True
        if object_id:
             extra_context['statement_url'] = reverse('select_vendor_statement_date', args=[object_id])
        return super().change_view(request, object_id, form_url, extra_context)


@admin.register(ExpenseCategory)
class ExpenseCategoryAdmin(RestrictedAdmin):
    list_display = ["name", "is_active"]
    search_fields = ["name"] 


@admin.register(Expense)
class ExpenseAdmin(RestrictedAdmin):
    actions = [export_expenses_detailed]
    list_display = [
        "date", "category", "description", "amount", "business_unit", 
        "is_paid_display", "payment_account", "vendor", "truck", "is_auto_synced", "recorded_by"
    ]
    list_filter = [
        "business_unit", "is_paid", "date", "category", "payment_account",
        "is_auto_synced", "vendor", "truck", "requires_approval"
    ]
    search_fields = ["description", "category__name", "vendor__name", "employee__name", "receipt_number", "notes"]
    date_hierarchy = "date"
    ordering = ["-date", "-created_at"]
    autocomplete_fields = ["category", "payment_account", "vendor", "truck", "machine", "employee", "approved_by"]
    readonly_fields = ["is_auto_synced", "business_unit"]

    fieldsets = (
        ("Expense Details", {
            "fields": ("date", "category", "description", "amount", "business_unit")
        }),
        ("Payment", {
            "fields": ("is_paid", "payment_date", "payment_account", "receipt_number"),
            "description": "Uncheck 'Is Paid' for credit purchases (pay vendor later)"
        }),
        ("Related To (Optional)", {
            "fields": ("vendor", "truck", "transport_asset", "machine", "employee", "driver"),
            "description": "Link expense to vendor, vehicle, machine, or employee. Business unit auto-sets based on truck/machine."
        }),
        ("Approval (For Large Expenses)", {
            "fields": ("requires_approval", "approved_by"),
            "classes": ("collapse",)
        }),
        ("System Info", {
            "fields": ("is_auto_synced",),
            "classes": ("collapse",),
            "description": "Auto-synced expenses are created automatically from Procurement/Maintenance"
        }),
        ("Notes", {"fields": ("notes",)}),
    )

    @admin.display(description="Paid?", boolean=True)
    def is_paid_display(self, obj):
        return obj.is_paid

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(ProcurementLog)
class ProcurementLogAdmin(RestrictedAdmin):
    actions = [export_procurement_detailed]
    list_display = [
        "date", "material", "quantity", "total_cost", "vendor",
        "is_paid_display", "payment_account", "is_internal_haulage", "haulage_fee", "delivery_truck"
    ]
    list_filter = ["date", "material", "vendor", "is_paid", "is_internal_haulage", "delivery_truck"]
    search_fields = ["material__name", "vendor__name", "remark"]
    date_hierarchy = "date"
    readonly_fields = ["unit_price"]
    ordering = ["-date", "-created_at"]
    autocomplete_fields = ["vendor", "delivery_truck", "payment_account"]

    fieldsets = (
        ("Purchase Info", {
            "fields": ("date", "material", "vendor")
        }),
        ("Quantity & Cost", {
            "fields": ("quantity", "total_cost", "unit_price")
        }),
        ("Payment", {
            "fields": ("is_paid", "payment_account"),
            "description": "Uncheck 'Is Paid' if vendor gave on credit. Vendor balance will update automatically."
        }),
        ("Internal Haulage (Transport Revenue)", {
            "fields": ("is_internal_haulage", "delivery_truck", "haulage_fee"),
            "description": "Check if Jafan Transport delivered. Haulage fee credits Transport."
        }),
        ("Notes", {
            "fields": ("remark",)
        }),
    )

    @admin.display(description="Paid?", boolean=True)
    def is_paid_display(self, obj):
        return obj.is_paid
    
    def save_model(self, request, obj, form, change):
        if not obj.recorded_by:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(ReturnLog)
class ReturnLogAdmin(RestrictedAdmin):
    actions = [export_returns_detailed]
    list_display = [
        "date", "customer", "site", "block_type", "quantity_returned",
        "condition", "restock_target", "credit_customer", "credit_value",
        "approved_by"
    ]
    list_filter = ["date", "condition", "credit_customer", "block_type", "customer", "restock_as"]
    search_fields = ["customer__name", "customer__phone", "site__name", "reason"]
    date_hierarchy = "date"
    ordering = ["-date", "-created_at"]
    autocomplete_fields = ["customer", "site", "original_supply", "block_type", "approved_by", "restock_as"]
    readonly_fields = ["credit_value"]

    fieldsets = (
        ("Return Details", {"fields": ("date", "customer", "site", "original_supply")}),
        ("Product Info", {"fields": ("block_type", "quantity_returned", "condition")}),
        ("Inventory Transformation", {
            "fields": ("restock_as",),
            "description": "Select ONLY if the returned block is now a different type (e.g., Whole → Half). Leave empty to restock as original type."
        }),
        ("Customer Credit", {
            "fields": ("credit_customer", "unit_price", "restocking_fee", "credit_value"),
            "description": "⚠️ IMPORTANT: Only check 'Credit Customer' if this is a genuine refund. "
                           "For excess blocks from delivery or breakage conversions, leave unchecked."
        }),
        ("Approval", {"fields": ("reason", "approved_by")}),
    )

    def restock_target(self, obj):
        if obj.condition == 'DAMAGED': return "❌ No Restock"
        target = obj.restock_as if obj.restock_as else obj.block_type
        return f"→ {target.name}"
    restock_target.short_description = "Restocked To"

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(CashRefund)
class CashRefundAdmin(RestrictedAdmin):
    actions = [export_to_excel]
    list_display = ["date", "customer", "amount", "payment_account", "approved_by", "recorded_by"]
    list_filter = ["date", "payment_account", "customer"]
    search_fields = ["customer__name", "customer__phone", "reason"]
    date_hierarchy = "date"
    ordering = ["-date", "-created_at"]
    autocomplete_fields = ["customer", "payment_account", "approved_by"]

    fieldsets = (
        ("Refund Details", {"fields": ("date", "customer", "amount")}),
        ("Payment", {"fields": ("payment_account",)}),
        ("Approval", {"fields": ("reason", "approved_by")}),
    )

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(BreakageLog)
class BreakageLogAdmin(RestrictedAdmin):
    list_display = [
        "date", "block_type", "quantity_broken", "reason",
        "convert_to_half", "half_block_type", "quantity_salvaged",
        "approved_by", "recorded_by"
    ]
    list_filter = ["date", "reason", "convert_to_half", "block_type", "approved_by"]
    search_fields = ["block_type__name", "description"]
    date_hierarchy = "date"
    ordering = ["-date", "-created_at"]
    autocomplete_fields = ["block_type", "half_block_type", "approved_by"]
    readonly_fields = ["recorded_by"]

    fieldsets = (
        ("Breakage Details", {
            "fields": ("date", "block_type", "quantity_broken")
        }),
        ("Reason", {
            "fields": ("reason", "description")
        }),
        ("Half Block Conversion", {
            "fields": ("convert_to_half", "half_block_type", "quantity_salvaged"),
            "description": "If broken blocks can be salvaged as half blocks, check 'Convert to half' and select the target half block type. Quantity salvaged is auto-suggested as 2x broken quantity."
        }),
        ("Approval", {
            "fields": ("approved_by",)
        }),
    )

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(FuelLog)
class FuelLogAdmin(RestrictedAdmin):
    list_display = [
        "date", "fuel_type", "destination_display", "driver",
        "quantity", "cost_per_liter", "total_cost", "is_paid", "payment_account", "dispensed_by"
    ]
    list_filter = ["date", "fuel_type", "destination_type", "truck", "machine", "transport_asset", "is_paid", "payment_method"]
    search_fields = ["truck__name", "machine__name", "transport_asset__name", "driver__name", "fuel_station", "remark"]
    date_hierarchy = "date"
    autocomplete_fields = ["truck", "machine", "driver", "payment_account"]
    readonly_fields = ["total_cost", "dispensed_by"]

    fieldsets = (
        ("Dispense Details", {
            "fields": ("date", "fuel_type", "destination_type", "truck", "machine", "transport_asset", "driver")
        }),
        ("Volume & Cost", {
            "fields": ("quantity", "cost_per_liter", "total_cost"),
            "description": "Cost per liter auto-fills from Material price for Diesel, but you can override for Petrol/Oil."
        }),
        ("Payment", {
            "fields": ("fuel_station", "payment_method", "is_paid", "payment_account")
        }),
        ("Metrics (Optional)", {
            "fields": ("engine_hours",)
        }),
        ("Notes", {
            "fields": ("remark", "dispensed_by")
        }),
    )

    def destination_display(self, obj):
        if obj.truck: return f"🚛 {obj.truck.name}"
        if obj.machine: return f"⚙️ {obj.machine.name}"
        if obj.transport_asset: return f"🏍️ {obj.transport_asset.name}"
        return "-"
    destination_display.short_description = "Destination"

    def save_model(self, request, obj, form, change):
        if not obj.dispensed_by:
            obj.dispensed_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(MaintenanceLog)
class MaintenanceLogAdmin(RestrictedAdmin):
    list_display = [
        "date", "target_display", "service_type", "cost",
        "payment_account", "vendor", "next_service_date", "recorded_by"
    ]
    list_filter = ["date", "target_type", "service_type", "truck", "transport_asset", "payment_account"]
    search_fields = ["truck__name", "machine__name", "transport_asset__name", "description", "vendor__name"]
    autocomplete_fields = ["truck", "machine", "vendor", "payment_account"]
    readonly_fields = ["expense_entry", "recorded_by"]

    fieldsets = (
        ("Asset Info", {
            "fields": ("date", "target_type", "truck", "machine", "transport_asset")
        }),
        ("Service Details", {
            "fields": ("service_type", "description", "vendor", "cost", "payment_account")
        }),
        ("Follow Up", {
            "fields": ("next_service_date",)
        }),
        ("System Info", {
            "fields": ("expense_entry", "recorded_by"),
            "classes": ("collapse",),
            "description": "Expense entry is auto-created for financial tracking."
        }),
    )

    def target_display(self, obj):
        if obj.truck: return f"🚛 {obj.truck.name}"
        if obj.machine: return f"⚙️ {obj.machine.name}"
        if obj.transport_asset: return f"🏍️ {obj.transport_asset.name}"
        return "-"
    target_display.short_description = "Asset"

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by: obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(Truck)
class TruckAdmin(RestrictedAdmin):
    actions = [export_to_excel]
    list_display = ["name", "truck_type", "plate_number", "driver", "status", "benchmark_fuel", "expected_trips", "is_active"]
    list_filter = ["truck_type", "status", "fuel_type", "is_active"]
    search_fields = ["name", "plate_number", "driver__name"]
    ordering = ["name"]
    autocomplete_fields = ["driver"]

    fieldsets = (
        ("Vehicle Info", {"fields": ("name", "truck_type", "plate_number", "driver", "status", "is_active")}),
        ("Fuel Benchmarks", {"fields": ("fuel_type", "fuel_capacity", "benchmark_fuel", "expected_trips")}),
    )


@admin.register(TransportAsset)
class TransportAssetAdmin(RestrictedAdmin):
    list_display = ["name", "asset_type", "plate_number", "assigned_to", "fuel_type", "is_active"]
    list_filter = ['asset_type', 'fuel_type', 'is_active']
    search_fields = ["name", "plate_number"] 
    list_editable = ['is_active']

    fieldsets = (
        (None, {
            'fields': ('name', 'asset_type', 'plate_number')
        }),
        ('Assignment', {
            'fields': ('assigned_to', 'fuel_type')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )


@admin.register(TransportRevenue)
class TransportRevenueAdmin(RestrictedAdmin):
    list_display = ['date', 'job_type', 'truck', 'driver', 'customer_name', 'trips', 'amount', 'is_paid', 'payment_method', 'payment_account']
    list_filter = ['job_type', 'is_paid', 'payment_method', 'truck', 'date']
    search_fields = ['customer_name', 'customer_phone', 'delivery_address', 'description']
    date_hierarchy = 'date'
    list_editable = ['is_paid']
    autocomplete_fields = ['truck', 'driver', 'payment_account']

    fieldsets = (
        ('Job Details', {
            'fields': ('date', 'job_type', 'truck', 'driver')
        }),
        ('Customer', {
            'fields': ('customer_name', 'customer_phone', 'delivery_address')
        }),
        ('Billing', {
            'fields': ('trips', 'amount', 'is_paid', 'payment_method', 'payment_account')
        }),
        ('Notes', {
            'fields': ('description',),
            'classes': ('collapse',)
        }),
    )

    def save_model(self, request, obj, form, change):
        if not change:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(BankCharge)
class BankChargeAdmin(RestrictedAdmin):
    list_display = ['date', 'account', 'charge_type', 'amount', 'description', 'reference']
    list_filter = ['charge_type', 'account', 'date']
    search_fields = ['description', 'reference']
    date_hierarchy = 'date'

    fieldsets = (
        (None, {
            'fields': ('date', 'account', 'charge_type', 'amount')
        }),
        ('Details', {
            'fields': ('description', 'reference')
        }),
    )

    def save_model(self, request, obj, form, change):
        if not change:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(AccountTransfer)
class AccountTransferAdmin(RestrictedAdmin):
    list_display = ["date", "from_account", "to_account", "amount", "is_transport_settlement", "reference"]
    list_filter = ['is_transport_settlement', 'from_account', 'to_account', 'date']
    search_fields = ['reference', 'description']
    date_hierarchy = 'date'
    readonly_fields = ["expense_entry"]

    fieldsets = (
        ('Transfer Details', {
            'fields': ('date', 'from_account', 'to_account', 'amount')
        }),
        ('Reference', {
            'fields': ('reference', 'description')
        }),
        ('Transport Settlement', {
            'fields': ('is_transport_settlement',),
            'description': 'Check this if Block Industry is paying Transport for deliveries (Monday settlement)'
        }),
    )

    def save_model(self, request, obj, form, change):
        if not change:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)
    
    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        transport_vendor = Vendor.objects.filter(is_internal=True, supply_type='TRANSPORT').first()
        if transport_vendor:
            extra_context['transport_outstanding'] = transport_vendor.account_balance
        else:
            extra_context['transport_outstanding'] = 0
        return super().changeform_view(request, object_id, form_url, extra_context)
    

@admin.register(VendorPayment)
class VendorPaymentAdmin(RestrictedAdmin):
    list_display = ['date', 'vendor', 'amount_display', 'payment_account', 'reference', 'recorded_by']
    list_filter = ['date', 'vendor', 'payment_account']
    search_fields = ['vendor__name', 'reference', 'description']
    date_hierarchy = 'date'
    ordering = ['-date', '-created_at']
    autocomplete_fields = ['vendor', 'payment_account']

    fieldsets = (
        ('Payment Details', {
            'fields': ('date', 'vendor', 'amount'),
            'description': 'Record payment TO a vendor to settle outstanding balance.'
        }),
        ('Payment Method', {
            'fields': ('payment_account', 'reference'),
            'description': 'Select the account you are paying FROM.'
        }),
        ('Notes', {
            'fields': ('description',)
        }),
    )

    @admin.display(description="Amount")
    def amount_display(self, obj):
        return format_html('<span style="color: #dc3545; font-weight: bold;">₦{}</span>', f"{obj.amount:,.2f}")

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        if object_id:
            obj = VendorPayment.objects.get(pk=object_id)
            extra_context['vendor_balance'] = obj.vendor.account_balance
        return super().changeform_view(request, object_id, form_url, extra_context)



@admin.register(TeamPayment)
class TeamPaymentAdmin(RestrictedAdmin):
    list_display = [
        'date', 'payment_type', 'team', 'employee', 'period_display',
        'calculated_amount_display', 'amount_paid_display', 'difference_display',
        'payment_account', 'recorded_by'
    ]
    list_filter = ['date', 'payment_type', 'team', 'payment_account']
    search_fields = ['team__name', 'employee__name', 'reference', 'description']
    date_hierarchy = 'date'
    ordering = ['-date', '-created_at']
    autocomplete_fields = ['team', 'employee', 'payment_account']
    readonly_fields = ['calculated_amount']

    fieldsets = (
        ('Payment Details', {
            'fields': ('date', 'payment_type', 'team', 'employee'),
            'description': 'Select team OR employee (not both)'
        }),
        ('Work Period (Optional)', {
            'fields': ('period_start', 'period_end'),
            'description': 'Only needed for Team Production Pay to auto-calculate from ProductionLog',
            'classes': ('collapse',)
        }),
        ('Amounts', {
            'fields': ('calculated_amount', 'amount_paid'),
            'description': 'Calculated amount auto-fills for Team Pay when period is set'
        }),
        ('Payment Method', {
            'fields': ('payment_account', 'reference')
        }),
        ('Notes', {
            'fields': ('description',),
            'description': "E.g., 'Daily loading - 500 blocks' or 'Stacking at ABC site'"
        }),
    )

    @admin.display(description="Period")
    def period_display(self, obj):
        if obj.period_start and obj.period_end:
            return f"{obj.period_start} to {obj.period_end}"
        return "-"

    @admin.display(description="Calculated")
    def calculated_amount_display(self, obj):
        if obj.calculated_amount > 0:
            return f"₦{obj.calculated_amount:,.2f}"
        return "-"

    @admin.display(description="Paid")
    def amount_paid_display(self, obj):
        return format_html(
            '<span style="color: #dc3545; font-weight: bold;">₦{}</span>',
            f"{obj.amount_paid:,.2f}"
        )

    @admin.display(description="Diff")
    def difference_display(self, obj):
        if obj.calculated_amount == 0:
            return "-"
        diff = obj.amount_paid - obj.calculated_amount
        if diff == 0:
            return format_html('<span style="color: green;">✓</span>')
        elif diff > 0:
            return format_html('<span style="color: orange;">+₦{}</span>', f"{diff:,.2f}")
        else:
            return format_html('<span style="color: blue;">-₦{}</span>', f"{abs(diff):,.2f}")

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(SandVehicleType)
class SandVehicleTypeAdmin(RestrictedAdmin):
    list_display = ['name', 'price_display', 'description', 'is_active']
    list_filter = ['is_active']
    search_fields = ['name', 'description']
    ordering = ['name']

    fieldsets = (
        (None, {
            'fields': ('name', 'price', 'description', 'is_active')
        }),
    )

    @admin.display(description="Price")
    def price_display(self, obj):
        return f"₦{obj.price:,.2f}"


@admin.register(SandSale)
class SandSaleAdmin(RestrictedAdmin):
    list_display = [
        'date', 'vehicle_type', 'quantity', 'total_amount_display',
        'payment_account', 'customer_name', 'recorded_by', 'pdf_actions'
    ]
    list_filter = ['date', 'vehicle_type', 'payment_account']
    search_fields = ['customer_name', 'customer_phone', 'remark']
    date_hierarchy = 'date'
    ordering = ['-date', '-created_at']
    autocomplete_fields = ['vehicle_type', 'payment_account']
    readonly_fields = ['unit_price', 'total_amount']

    fieldsets = (
        ('Sale Details', {
            'fields': ('date', 'vehicle_type', 'quantity')
        }),
        ('Payment', {
            'fields': ('payment_account', 'unit_price', 'total_amount')
        }),
        ('Customer Info (Optional)', {
            'fields': ('customer_name', 'customer_phone'),
            'classes': ('collapse',)
        }),
        ('Notes', {
            'fields': ('remark',)
        }),
    )

    @admin.display(description="Total")
    def total_amount_display(self, obj):
        return format_html(
            '<span style="color: #28a745; font-weight: bold;">₦{}</span>',
            f"{obj.total_amount:,.2f}"
        )

    def pdf_actions(self, obj):
        receipt_url = reverse('generate_sand_receipt', args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" style="padding:3px 8px; background:#2E7D32; color:white; border-radius:4px; text-decoration:none; font-size:11px;">🧾 Receipt</a>',
            receipt_url
        )
    pdf_actions.short_description = "PDF"

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(Debtor)
class DebtorAdmin(RestrictedAdmin):
    list_display = ['name', 'phone', 'employee_link', 'loan_balance_display', 'balance_status', 'is_active']
    list_filter = ['is_active', 'employee']
    search_fields = ['name', 'phone', 'employee__name']
    ordering = ['name']
    autocomplete_fields = ['employee']

    fieldsets = (
        ('Personal Info', {
            'fields': ('name', 'phone', 'employee', 'address', 'id_number')
        }),
        ('Balance', {
            'fields': ('loan_balance',),
            'description': 'Positive = They owe us'
        }),
        ('Status', {
            'fields': ('is_active', 'notes')
        }),
    )
    readonly_fields = ['loan_balance']

    @admin.display(description="Employee")
    def employee_link(self, obj):
        if obj.employee:
            return format_html('<span style="color: green;">✓ {}</span>', obj.employee.name)
        return mark_safe('<span style="color: gray;">External</span>')

    @admin.display(description="Balance")
    def loan_balance_display(self, obj):
        if obj.loan_balance > 0:
            return format_html('<span style="color: #dc3545; font-weight: bold;">₦{}</span>', f"{obj.loan_balance:,.2f}")
        elif obj.loan_balance < 0:
            return format_html('<span style="color: #28a745;">₦{}</span>', f"{obj.loan_balance:,.2f}")
        return "₦0.00"


@admin.register(Loan)
class LoanAdmin(RestrictedAdmin):
    list_display = [
        'loan_id', 'date', 'debtor', 'amount_display', 'repaid_display',
        'outstanding_display', 'progress_display', 'repayment_mode', 
        'is_fully_repaid', 'approved_by', 'pdf_actions'
    ]
    list_filter = ['date', 'repayment_mode', 'is_fully_repaid', 'payment_account', 'approved_by']
    search_fields = ['debtor__name', 'debtor__phone', 'purpose', 'reference']
    date_hierarchy = 'date'
    ordering = ['-date', '-created_at']
    autocomplete_fields = ['debtor', 'payment_account', 'approved_by']
    readonly_fields = ['amount_repaid', 'is_fully_repaid']

    fieldsets = (
        ('Loan Details', {
            'fields': ('date', 'debtor', 'amount', 'payment_account')
        }),
        ('Repayment Terms', {
            'fields': ('purpose', 'repayment_mode', 'expected_repayment_date')
        }),
        ('Tracking', {
            'fields': ('amount_repaid', 'is_fully_repaid'),
            'classes': ('collapse',)
        }),
        ('Approval', {
            'fields': ('approved_by', 'reference', 'notes'),
            'description': 'Only Administrators and General Managers can approve loans.'
        }),
    )

    @admin.display(description="Loan ID")
    def loan_id(self, obj):
        return f"LOAN-{obj.pk:05d}"

    @admin.display(description="Amount")
    def amount_display(self, obj):
        return format_html('<span style="font-weight: bold;">₦{}</span>', f"{obj.amount:,.2f}")

    @admin.display(description="Repaid")
    def repaid_display(self, obj):
        if obj.amount_repaid > 0:
            return format_html('<span style="color: #28a745;">₦{}</span>', f"{obj.amount_repaid:,.2f}")
        return "₦0.00"

    @admin.display(description="Outstanding")
    def outstanding_display(self, obj):
        outstanding = obj.outstanding_balance
        if outstanding > 0:
            return format_html('<span style="color: #dc3545; font-weight: bold;">₦{}</span>', f"{outstanding:,.2f}")
        return mark_safe('<span style="color: #28a745;">✓ Cleared</span>')

    @admin.display(description="Progress")
    def progress_display(self, obj):
        p = obj.repayment_progress
        if p >= 100:
            color = "green"
        elif p > 50:
            color = "orange"
        else:
            color = "gray"
        return format_html('<span style="color: {}; font-weight: bold;">{}%</span>', color, p)

    def pdf_actions(self, obj):
        statement_url = reverse('generate_loan_statement', args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" style="padding:3px 8px; background:#2E7D32; color:white; border-radius:4px; text-decoration:none; font-size:11px;">📄 Statement</a>',
            statement_url
        )
    pdf_actions.short_description = "PDF"

    def get_readonly_fields(self, request, obj=None):
        """Make approved_by readonly for non-approvers, or if already approved."""
        readonly = list(super().get_readonly_fields(request, obj) or [])
        
        role = self._get_user_role(request.user)
        
        # If user cannot approve loans, make approved_by readonly
        if role not in self.LOAN_APPROVAL_ROLES:
            if 'approved_by' not in readonly:
                readonly.append('approved_by')
        
        # If loan is already approved and user is not admin, lock the field
        if obj and obj.approved_by and role != 'ADMIN':
            if 'approved_by' not in readonly:
                readonly.append('approved_by')
        
        return readonly

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Limit approved_by choices to users who can approve (ADMIN and MANAGER only)."""
        if db_field.name == 'approved_by':
            from .models import User
            approver_ids = []
            for user in User.objects.filter(is_active=True, is_staff=True):
                if user.is_superuser:
                    approver_ids.append(user.pk)
                elif hasattr(user, 'employee_profile') and user.employee_profile:
                    if user.employee_profile.role in ['ADMIN', 'MANAGER']:
                        approver_ids.append(user.pk)
            kwargs['queryset'] = User.objects.filter(pk__in=approver_ids)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(LoanRepayment)
class LoanRepaymentAdmin(RestrictedAdmin):
    list_display = ['date', 'loan_display', 'debtor_name', 'amount_display', 'repayment_method', 'payment_account', 'recorded_by']
    list_filter = ['date', 'repayment_method', 'payment_account', 'loan__debtor']
    search_fields = ['loan__debtor__name', 'reference', 'notes']
    date_hierarchy = 'date'
    ordering = ['-date', '-created_at']
    autocomplete_fields = ['loan', 'payment_account']

    fieldsets = (
        ('Repayment Details', {
            'fields': ('date', 'loan', 'amount')
        }),
        ('Payment Info', {
            'fields': ('payment_account', 'repayment_method', 'reference')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )

    @admin.display(description="Loan")
    def loan_display(self, obj):
        return f"LOAN-{obj.loan.pk:05d}"

    @admin.display(description="Debtor")
    def debtor_name(self, obj):
        return obj.loan.debtor.name

    @admin.display(description="Amount")
    def amount_display(self, obj):
        return format_html('<span style="color: #28a745; font-weight: bold;">₦{}</span>', f"{obj.amount:,.2f}")

    def save_model(self, request, obj, form, change):
        if not obj.recorded_by:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(QuickSale)
class QuickSaleAdmin(RestrictedAdmin):
    list_display = ['quick_sale_id', 'date', 'block_type', 'quantity', 'unit_price_display', 'total_amount_display', 'payment_summary', 'buyer_name', 'recorded_by', 'receipt_link']
    list_filter = ['date', 'block_type', 'payment_method', 'payment_account']
    search_fields = ['buyer_name', 'buyer_phone', 'reference']
    date_hierarchy = 'date'
    autocomplete_fields = ['block_type', 'payment_account', 'secondary_account']
    readonly_fields = ['unit_price', 'total_amount', 'primary_amount_display', 'recorded_by', 'created_at']
    
    fieldsets = (
        ('Sale Details', {
            'fields': ('date', 'block_type', 'quantity', 'unit_price', 'logistics_discount', 'total_amount')
        }),
        ('Primary Payment', {
            'fields': ('payment_account', 'payment_method', 'reference', 'primary_amount_display')
        }),
        ('Split Payment (Optional)', {
            'fields': ('secondary_amount', 'secondary_account', 'secondary_method', 'secondary_reference'),
            'classes': ('collapse',),
            'description': 'Use this section if customer pays with two methods (e.g., part Cash, part POS)'
        }),
        ('Buyer Info (Optional)', {
            'fields': ('buyer_name', 'buyer_phone', 'pickup_authorized_by'),
            'classes': ('collapse',)
        }),
        ('Notes', {
            'fields': ('remark',),
            'classes': ('collapse',)
        }),
        ('System', {
            'fields': ('recorded_by', 'created_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_readonly_fields(self, request, obj=None):
        """Lock financial fields after creation to prevent balance corruption."""
        readonly = list(super().get_readonly_fields(request, obj))
        
        if obj:  # Editing existing record
            readonly.extend([
                'block_type', 'quantity', 'logistics_discount',
                'payment_account', 'payment_method',
                'secondary_amount', 'secondary_account', 'secondary_method'
            ])
        
        return readonly
    
    def quick_sale_id(self, obj):
        return f"QS-{obj.pk:05d}"
    quick_sale_id.short_description = "Sale ID"
    
    def unit_price_display(self, obj):
        return f"₦{obj.unit_price:,.2f}"
    unit_price_display.short_description = "Unit Price"
    
    def total_amount_display(self, obj):
        return f"₦{obj.total_amount:,.2f}"
    total_amount_display.short_description = "Total"
    
    def primary_amount_display(self, obj):
        if obj.pk:
            return f"₦{obj.primary_amount:,.2f}"
        return "Calculated on save"
    primary_amount_display.short_description = "Primary Amount"
    
    def payment_summary(self, obj):
        if obj.is_split_payment:
            return f"{obj.payment_method} + {obj.secondary_method}"
        return obj.payment_method
    payment_summary.short_description = "Payment"
    
    def receipt_link(self, obj):
        url = reverse('generate_quick_sale_receipt', args=[obj.pk])
        return format_html('<a href="{}" target="_blank">📄 Receipt</a>', url)
    receipt_link.short_description = "Receipt"
    
    def save_model(self, request, obj, form, change):
        if not change:
            obj.recorded_by = request.user
        super().save_model(request, obj, form, change)