from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db.models import Sum, Q, F
from decimal import Decimal

from django.http import JsonResponse


from .kpi_service import KPIService
from .services import BlockIndustryPLService, CashFlowService
from .exports import ReportExporter
from .models import (
    SupplyLog, Payment, Customer, SalesOrder, 
    Vendor, PaymentAccount, Expense, FuelLog, MaintenanceLog,
    Truck, TransportRevenue, TransportAsset, SalesOrderItem, SandSale, Loan, LoanRepayment, Debtor,


)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .pdf_generator import (
    InvoiceGenerator,
    WaybillGenerator,
    ReceiptGenerator,
    CustomerStatementGenerator,
    ProformaInvoiceGenerator,
    VendorStatementGenerator,
    AccountStatementGenerator,
    ProfitLossGenerator,
    CashFlowPDFGenerator,
    SandSaleReceiptGenerator,
    LoanStatementGenerator,
    LoanReportGenerator
    

)


# ==================== DASHBOARD VIEWS ====================

@staff_member_required
def dashboard_view(request):
    """
    Executive Dashboard with Accurate P&L.
    Shows true profitability using industry-standard COGS calculation.
    """
    today = timezone.now().date()
    
    # Date range from request or default to current month
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = today.replace(day=1)  # First of current month
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = today
    
    # Get accurate P&L data
    pl_service = BlockIndustryPLService(start_date, end_date, business_unit='BLOCK')
    pl_data = pl_service.get_full_pl()
    
    # Legacy KPI service for other stats
    kpi_service = KPIService()
    
    # Cash position
    block_accounts = PaymentAccount.objects.filter(business_unit='BLOCK', is_active=True)
    cash_position = block_accounts.aggregate(total=Sum('current_balance'))['total'] or Decimal('0')
    
    # Receivables (customers who owe us - negative balance means they owe)
    total_receivables = Customer.objects.filter(
        account_balance__lt=0
    ).aggregate(total=Sum('account_balance'))['total'] or Decimal('0')
    total_receivables = abs(total_receivables)
    
    # Payables (we owe vendors - positive balance means we owe)
    total_payables = Vendor.objects.filter(
        account_balance__gt=0,
        is_internal=False
    ).aggregate(total=Sum('account_balance'))['total'] or Decimal('0')
    
    context = {
        "title": "Executive Dashboard",
        "site_header": "Jafan ERP",
        
        # Period
        "start_date": start_date,
        "end_date": end_date,
        
        # P&L Summary
        "revenue": pl_data['revenue']['total'],
        "revenue_breakdown": pl_data['revenue']['by_block_type'],
        "blocks_sold": pl_data['revenue']['total_quantity_sold'],
        "logistics_income": pl_data['revenue']['logistics_income'],
        
        "cogs": pl_data['cogs']['total'],
        "cogs_breakdown": pl_data['cogs']['by_block_type'],
        
        "gross_profit": pl_data['gross_profit'],
        "gross_margin": pl_data['gross_margin_percent'],
        
        "operating_expenses": pl_data['operating_expenses']['total'],
        "expenses_breakdown": pl_data['operating_expenses']['breakdown'],
        
        "net_profit": pl_data['net_profit'],
        "net_margin": pl_data['net_margin_percent'],
        
        # Balance Sheet Items
        "cash_position": cash_position,
        "receivables": total_receivables,
        "payables": total_payables,
        
        # Legacy stats
        "debtors": kpi_service.get_top_debtors(),
        "alerts": kpi_service.get_inventory_alerts(),
        "activity": kpi_service.get_recent_activity(),
    }
    
    return render(request, "admin/erp/dashboard.html", context)


@staff_member_required
def transport_dashboard_view(request):
    """
    Jafan Transport Dashboard.
    Tracks Revenue, Costs, Profit, and Diesel Efficiency.
    """
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())  # Monday
    
    # 1. ACCOUNTS & BALANCES
    transport_vendor = Vendor.objects.filter(is_internal=True, supply_type='TRANSPORT').first()
    receivable_from_block = transport_vendor.account_balance if transport_vendor else Decimal('0')
    
    transport_accounts = PaymentAccount.objects.filter(business_unit='TRANSPORT', is_active=True)
    total_transport_balance = transport_accounts.aggregate(total=Sum('current_balance'))['total'] or Decimal('0')
    
    # 2. INCOME (This Week)
    week_logistics_income = SupplyLog.objects.filter(
        date__gte=week_start,
        delivery_type='DELIVERED'
    ).aggregate(total=Sum('logistics_income'))['total'] or Decimal('0')
    
    week_external_income = TransportRevenue.objects.filter(
        date__gte=week_start
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    total_income = week_logistics_income + week_external_income
    
    # 3. EXPENSES (This Week)
    week_expenses = Expense.objects.filter(
        date__gte=week_start,
        business_unit='TRANSPORT',
        is_paid=True
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Breakout for display
    week_fuel_cost = FuelLog.objects.filter(
        date__gte=week_start,
        destination_type__in=['TRUCK', 'ASSET']
    ).aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    
    week_maintenance_cost = MaintenanceLog.objects.filter(
        date__gte=week_start,
        target_type__in=['TRUCK', 'ASSET']
    ).aggregate(total=Sum('cost'))['total'] or Decimal('0')
    
    net_profit = total_income - week_expenses

    # 4. DIESEL EFFICIENCY AUDIT (Theft Check)
    actual_fuel_liters = FuelLog.objects.filter(
        date__gte=week_start, 
        destination_type='TRUCK',
        fuel_type='DIESEL'
    ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')

    expected_fuel_liters = Decimal('0')
    active_trucks = Truck.objects.filter(is_active=True)
    supplies_week = SupplyLog.objects.filter(date__gte=week_start, delivery_type='DELIVERED')
    outside_trips = 0

    for truck in active_trucks:
        trips_count = supplies_week.filter(truck=truck).count()
        if truck.expected_trips > 0:
            consumption_rate = truck.benchmark_fuel / truck.expected_trips
            expected_fuel_liters += (Decimal(trips_count) * consumption_rate)
    
    for log in supplies_week:
        if log.site and log.site.is_outside_town:
            outside_trips += 1

    fuel_audit = {
        'actual_liters': actual_fuel_liters,
        'expected_liters': expected_fuel_liters,
        'variance': actual_fuel_liters - expected_fuel_liters,
        'outside_trips': outside_trips
    }

    # 5. PER-TRUCK BREAKDOWN
    truck_breakdown = []
    for truck in active_trucks:
        logistics = SupplyLog.objects.filter(truck=truck, date__gte=week_start, delivery_type='DELIVERED').aggregate(t=Sum('logistics_income'))['t'] or Decimal('0')
        external = TransportRevenue.objects.filter(truck=truck, date__gte=week_start).aggregate(t=Sum('amount'))['t'] or Decimal('0')
        
        direct_fuel = FuelLog.objects.filter(truck=truck, date__gte=week_start).aggregate(t=Sum('total_cost'))['t'] or Decimal('0')
        direct_maint = MaintenanceLog.objects.filter(truck=truck, date__gte=week_start).aggregate(t=Sum('cost'))['t'] or Decimal('0')
        other_exp = Expense.objects.filter(truck=truck, date__gte=week_start).aggregate(t=Sum('amount'))['t'] or Decimal('0')
        
        total_truck_income = logistics + external
        truck_net = total_truck_income - other_exp 

        truck_breakdown.append({
            'name': truck.name,
            'driver': truck.driver.name if truck.driver else 'Unassigned',
            'logistics_income': logistics,
            'external_income': external,
            'expenses': other_exp,
            'fuel': direct_fuel,
            'maintenance': direct_maint,
            'net_profit': truck_net
        })
    
    context = {
        'title': 'Transport Dashboard',
        'site_header': 'Jafan ERP',
        'transport_accounts': transport_accounts,
        'total_transport_balance': total_transport_balance,
        'receivable_from_block': receivable_from_block,
        
        'week_start': week_start,
        'week_logistics_income': week_logistics_income,
        'week_external_income': week_external_income,
        'week_expenses': week_expenses,
        'week_fuel': week_fuel_cost,
        'week_maintenance': week_maintenance_cost,
        'total_income': total_income,
        'total_costs': week_expenses,
        'net_profit': net_profit,
        
        'fuel_audit': fuel_audit,
        'truck_breakdown': truck_breakdown,
    }
    
    return render(request, 'admin/erp/transport_dashboard.html', context)


# ==================== CASH FLOW VIEW ====================

@staff_member_required
def cash_flow_view(request):
    """
    Cash Flow Statement View.
    Shows all money movements with filtering options.
    """
    today = timezone.now().date()
    
    # Parse date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    account_id = request.GET.get('account')
    business_unit = request.GET.get('business_unit')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = today.replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = today
    
    # Get account if specified
    account = None
    if account_id:
        account = PaymentAccount.objects.filter(pk=account_id).first()
    
    # Get cash flow data
    cf_service = CashFlowService(
        start_date, 
        end_date, 
        account=account,
        business_unit=business_unit if business_unit else None
    )
    cf_data = cf_service.get_cash_flow_statement()
    
    # Get all accounts for filter dropdown
    all_accounts = PaymentAccount.objects.filter(is_active=True).order_by('business_unit', 'bank_name')
    
    context = {
        'title': 'Cash Flow Statement',
        'site_header': 'Jafan ERP',
        
        # Filters
        'start_date': start_date,
        'end_date': end_date,
        'selected_account': account,
        'selected_business_unit': business_unit,
        'all_accounts': all_accounts,
        
        # Cash Flow Data
        'opening_balance': cf_data['opening_balance'],
        'closing_balance': cf_data['closing_balance'],
        'total_inflow': cf_data['total_inflow'],
        'total_outflow': cf_data['total_outflow'],
        'net_cash_flow': cf_data['net_cash_flow'],
        
        'transactions': cf_data['transactions'],
        'inflow_by_category': cf_data['inflow_by_category'],
        'outflow_by_category': cf_data['outflow_by_category'],
    }
    
    return render(request, 'admin/erp/cash_flow.html', context)


# ==================== P&L REPORT VIEW ====================

@staff_member_required
def pl_report_view(request):
    """
    Detailed Profit & Loss Report View.
    Shows full P&L breakdown with option to export.
    """
    today = timezone.now().date()
    
    # Parse date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = today.replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = today
    
    # Get P&L data
    pl_service = BlockIndustryPLService(start_date, end_date, business_unit='BLOCK')
    pl_data = pl_service.get_full_pl()
    
    context = {
        'title': 'Profit & Loss Statement',
        'site_header': 'Jafan ERP',
        
        # Period
        'start_date': start_date,
        'end_date': end_date,
        
        # Full P&L Data
        'pl_data': pl_data,
        
        # Unpacked for easy template access
        'revenue': pl_data['revenue'],
        'cogs': pl_data['cogs'],
        'gross_profit': pl_data['gross_profit'],
        'gross_margin': pl_data['gross_margin_percent'],
        'expenses': pl_data['operating_expenses'],
        'net_profit': pl_data['net_profit'],
        'net_margin': pl_data['net_margin_percent'],
    }
    
    return render(request, 'admin/erp/pl_report.html', context)


@staff_member_required
def generate_sand_receipt(request, sale_id):
    sale = get_object_or_404(SandSale, pk=sale_id)
    return SandSaleReceiptGenerator().generate(sale)

# ==================== DATE SELECTION HELPERS ====================

def parse_dates(request):
    """Helper to parse date filters from request."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    return start_date, end_date

@staff_member_required
def generate_loan_statement(request, loan_id):
    loan = get_object_or_404(Loan, pk=loan_id)
    return LoanStatementGenerator().generate(loan)


@staff_member_required
def loan_report_view(request):
    """View to select date range for loan reports."""
    today = timezone.now().date()
    
    if request.method == 'POST':
        start = request.POST.get('start_date')
        end = request.POST.get('end_date')
        export_type = request.POST.get('export_type', 'pdf')
        
        if start and end:
            if export_type == 'excel':
                return redirect(f"{reverse('export_loans_excel')}?start_date={start}&end_date={end}")
            else:
                return redirect(f"{reverse('export_loans_pdf')}?start_date={start}&end_date={end}")
        messages.error(request, "Please select both dates.")
    
    context = {
        'today': today.strftime('%Y-%m-%d'),
        'default_start': today.replace(day=1).strftime('%Y-%m-%d'),
        'title': 'Loan Report'
    }
    return render(request, 'admin/erp/loan_report_select.html', context)


@staff_member_required
def export_loans_pdf(request):
    """Export all loans for a period to PDF."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
    
    return LoanReportGenerator().generate(start_date, end_date)


@staff_member_required
def export_loans_excel(request):
    """Export all loans for a period to Excel."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
    
    loans = Loan.objects.filter(date__gte=start_date, date__lte=end_date).order_by('date')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Loans Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="254451", end_color="254451", fill_type="solid")
    currency_format = '₦#,##0.00'
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "JAFAN STANDARD BLOCK INDUSTRY - LOAN REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Period: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Loan ID', 'Date', 'Debtor', 'Phone', 'Amount', 'Repaid', 'Outstanding', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    total_amount = Decimal('0')
    total_repaid = Decimal('0')
    total_outstanding = Decimal('0')
    
    for loan in loans:
        ws.cell(row=row, column=1, value=f"LOAN-{loan.pk:05d}")
        ws.cell(row=row, column=2, value=loan.date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=loan.debtor.name)
        ws.cell(row=row, column=4, value=loan.debtor.phone or '-')
        ws.cell(row=row, column=5, value=loan.amount).number_format = currency_format
        ws.cell(row=row, column=6, value=loan.amount_repaid).number_format = currency_format
        ws.cell(row=row, column=7, value=loan.outstanding_balance).number_format = currency_format
        ws.cell(row=row, column=8, value="Cleared" if loan.is_fully_repaid else "Outstanding")
        
        total_amount += loan.amount
        total_repaid += loan.amount_repaid
        total_outstanding += loan.outstanding_balance
        row += 1
    
    # Totals
    row += 1
    ws.cell(row=row, column=4, value="TOTALS:").font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_amount).number_format = currency_format
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_repaid).number_format = currency_format
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_outstanding).number_format = currency_format
    ws.cell(row=row, column=7).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Loans_Report_{start_date}_{end_date}.xlsx'
    wb.save(response)
    return response


@staff_member_required
def select_statement_date(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)
    if request.method == 'POST':
        start = request.POST.get('start_date')
        end = request.POST.get('end_date')
        if start and end:
            url = reverse('generate_customer_statement', args=[customer_id])
            return redirect(f"{url}?start_date={start}&end_date={end}")
        messages.error(request, "Please select both dates.")

    context = {
        'customer': customer,
        'today': timezone.now().date().strftime('%Y-%m-%d'),
        'default_start': timezone.now().date().replace(day=1).strftime('%Y-%m-%d'),
        'title': f"Statement: {customer.name}"
    }
    return render(request, 'admin/erp/date_picker.html', context)


@staff_member_required
def select_vendor_statement_date(request, vendor_id):
    vendor = get_object_or_404(Vendor, pk=vendor_id)
    if request.method == 'POST':
        start = request.POST.get('start_date')
        end = request.POST.get('end_date')
        if start and end:
            url = reverse('generate_vendor_statement', args=[vendor_id])
            return redirect(f"{url}?start_date={start}&end_date={end}")
        messages.error(request, "Please select both dates.")

    context = {
        'vendor': vendor,
        'today': timezone.now().date().strftime('%Y-%m-%d'),
        'default_start': timezone.now().date().replace(day=1).strftime('%Y-%m-%d'),
        'title': f"Statement: {vendor.name}"
    }
    return render(request, 'admin/erp/date_picker_vendor.html', context)


@staff_member_required
def select_account_statement_date(request, account_id):
    account = get_object_or_404(PaymentAccount, pk=account_id)
    if request.method == 'POST':
        start = request.POST.get('start_date')
        end = request.POST.get('end_date')
        if start and end:
            url = reverse('generate_account_statement', args=[account_id])
            return redirect(f"{url}?start_date={start}&end_date={end}")
        messages.error(request, "Please select both dates.")

    context = {
        'account': account,
        'today': timezone.now().date().strftime('%Y-%m-%d'),
        'default_start': timezone.now().date().replace(day=1).strftime('%Y-%m-%d'),
        'title': f"Statement: {account.bank_name}"
    }
    return render(request, 'admin/erp/date_picker_account.html', context)


# ==================== EXCEL EXPORTS ====================

@staff_member_required
def export_sales_csv(request):
    s, e = parse_dates(request)
    return ReportExporter(s, e).export_sales_csv()

@staff_member_required
def export_sales_excel(request):
    s, e = parse_dates(request)
    return ReportExporter(s, e).export_sales_excel()

@staff_member_required
def export_expenses_csv(request):
    s, e = parse_dates(request)
    return ReportExporter(s, e).export_expenses_csv()

@staff_member_required
def export_expenses_excel(request):
    s, e = parse_dates(request)
    return ReportExporter(s, e).export_expenses_excel()

@staff_member_required
def export_production_csv(request):
    s, e = parse_dates(request)
    return ReportExporter(s, e).export_production_csv()

@staff_member_required
def export_production_excel(request):
    s, e = parse_dates(request)
    return ReportExporter(s, e).export_production_excel()

@staff_member_required
def export_customer_ledger(request):
    s, e = parse_dates(request)
    cid = request.GET.get('customer_id')
    return ReportExporter(s, e).export_customer_ledger_excel(cid)

@staff_member_required
def export_inventory(request):
    return ReportExporter().export_inventory_excel()


# ==================== PDF GENERATION ====================

@staff_member_required
def generate_invoice(request, supply_id):
    supply = get_object_or_404(SupplyLog, pk=supply_id)
    return InvoiceGenerator().generate(supply)

@staff_member_required
def generate_waybill(request, supply_id):
    supply = get_object_or_404(SupplyLog, pk=supply_id)
    return WaybillGenerator().generate(supply)

@staff_member_required
def generate_receipt(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id)
    return ReceiptGenerator().generate(payment)

@staff_member_required
def generate_proforma(request, order_id):
    order = get_object_or_404(SalesOrder, pk=order_id)
    return ProformaInvoiceGenerator().generate(order)

@staff_member_required
def generate_customer_statement(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)
    s, e = parse_dates(request)
    return CustomerStatementGenerator().generate(customer, s, e)

@staff_member_required
def generate_vendor_statement(request, vendor_id):
    vendor = get_object_or_404(Vendor, pk=vendor_id)
    s, e = parse_dates(request)
    return VendorStatementGenerator().generate(vendor, s, e)

@staff_member_required
def generate_account_statement(request, account_id):
    account = get_object_or_404(PaymentAccount, pk=account_id)
    s, e = parse_dates(request)
    return AccountStatementGenerator().generate(account, s, e)


# Add these to views.py
@staff_member_required
def generate_pl_pdf(request):
    """Generate P&L PDF report."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
    
    return ProfitLossGenerator().generate(start_date, end_date)


@staff_member_required
def generate_cashflow_pdf(request):
    """Generate Cash Flow PDF report."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    account_id = request.GET.get('account')
    business_unit = request.GET.get('business_unit')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
    
    account = None
    if account_id:
        account = PaymentAccount.objects.filter(pk=account_id).first()
    
    return CashFlowPDFGenerator().generate(start_date, end_date, account=account, business_unit=business_unit)


@staff_member_required
def export_pl_excel(request):
    """Export Profit & Loss to Excel."""
    from .services import BlockIndustryPLService
    
    # Parse dates
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
    
    # Get P&L data
    pl_service = BlockIndustryPLService(start_date, end_date, business_unit='BLOCK')
    pl_data = pl_service.get_full_pl()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="254451", end_color="254451", fill_type="solid")
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    bold_font = Font(bold=True)
    currency_format = '₦#,##0.00'
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "JAFAN STANDARD BLOCK INDUSTRY"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws['A2'] = "PROFIT & LOSS STATEMENT"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    ws['A3'] = f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}"
    ws['A3'].alignment = Alignment(horizontal='center')
    
    row = 5
    
    # REVENUE SECTION
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_fill
    ws[f'B{row}'].fill = light_fill
    ws[f'C{row}'].fill = light_fill
    row += 1
    
    for item in pl_data['revenue']['by_block_type']:
        ws[f'A{row}'] = f"  {item['block_type__name']}"
        ws[f'B{row}'] = f"{item['quantity'] or 0:,} blocks"
        ws[f'C{row}'] = item['revenue'] or 0
        ws[f'C{row}'].number_format = currency_format
        row += 1
    
    ws[f'A{row}'] = "Total Block Sales"
    ws[f'A{row}'].font = bold_font
    ws[f'C{row}'] = pl_data['revenue']['block_sales']
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = bold_font
    row += 1
    
    ws[f'A{row}'] = "  Logistics Income (to Transport)"
    ws[f'C{row}'] = pl_data['revenue']['logistics_income']
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = Font(italic=True, color="888888")
    row += 1
    
    ws[f'A{row}'] = "TOTAL REVENUE"
    ws[f'A{row}'].font = bold_font
    ws[f'C{row}'] = pl_data['revenue']['total']
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = bold_font
    row += 2
    
    # COGS SECTION
    ws[f'A{row}'] = "COST OF GOODS SOLD"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_fill
    ws[f'B{row}'].fill = light_fill
    ws[f'C{row}'].fill = light_fill
    row += 1
    
    for item in pl_data['cogs']['by_block_type']:
        ws[f'A{row}'] = f"  {item['block_type__name']}"
        ws[f'B{row}'] = f"{item['quantity'] or 0:,} blocks"
        ws[f'C{row}'] = item['cogs'] or 0
        ws[f'C{row}'].number_format = currency_format
        row += 1
    
    ws[f'A{row}'] = "TOTAL COGS"
    ws[f'A{row}'].font = bold_font
    ws[f'C{row}'] = pl_data['cogs']['total']
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = bold_font
    row += 2
    
    # GROSS PROFIT
    ws[f'A{row}'] = f"GROSS PROFIT ({pl_data['gross_margin_percent']}% margin)"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = gold_fill
    ws[f'B{row}'].fill = gold_fill
    ws[f'C{row}'] = pl_data['gross_profit']
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = bold_font
    ws[f'C{row}'].fill = gold_fill
    row += 2
    
    # OPERATING EXPENSES
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_fill
    ws[f'B{row}'].fill = light_fill
    ws[f'C{row}'].fill = light_fill
    row += 1
    
    for cat, amount in pl_data['operating_expenses']['breakdown'].items():
        ws[f'A{row}'] = f"  {cat}"
        ws[f'C{row}'] = amount
        ws[f'C{row}'].number_format = currency_format
        row += 1
    
    if not pl_data['operating_expenses']['breakdown']:
        ws[f'A{row}'] = "  (No operating expenses)"
        row += 1
    
    ws[f'A{row}'] = "TOTAL OPERATING EXPENSES"
    ws[f'A{row}'].font = bold_font
    ws[f'C{row}'] = pl_data['operating_expenses']['total']
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = bold_font
    row += 2
    
    # NET PROFIT
    ws[f'A{row}'] = f"NET PROFIT BEFORE TAX ({pl_data['net_margin_percent']}% margin)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'C{row}'] = pl_data['net_profit']
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = Font(bold=True, size=12, color="D4AF37")
    ws[f'C{row}'].fill = header_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ProfitLoss_{start_date}_{end_date}.xlsx'
    wb.save(response)
    return response


@staff_member_required
def export_cashflow_excel(request):
    """Export Cash Flow Statement to Excel."""
    from .services import CashFlowService
    
    # Parse parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    account_id = request.GET.get('account')
    business_unit = request.GET.get('business_unit')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
    
    account = None
    if account_id:
        account = PaymentAccount.objects.filter(pk=account_id).first()
    
    # Get Cash Flow data
    cf_service = CashFlowService(start_date, end_date, account=account, business_unit=business_unit)
    cf_data = cf_service.get_cash_flow_statement()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="254451", end_color="254451", fill_type="solid")
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    green_font = Font(color="28A745")
    red_font = Font(color="DC3545")
    bold_font = Font(bold=True)
    currency_format = '₦#,##0.00'
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "JAFAN STANDARD BLOCK INDUSTRY"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "CASH FLOW STATEMENT"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:H3')
    account_text = f" | Account: {account.bank_name}" if account else ""
    ws['A3'] = f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}{account_text}"
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Summary row
    row = 5
    ws[f'A{row}'] = "Opening Balance:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = cf_data['opening_balance']
    ws[f'B{row}'].number_format = currency_format
    
    ws[f'D{row}'] = "Total Inflow:"
    ws[f'D{row}'].font = bold_font
    ws[f'E{row}'] = cf_data['total_inflow']
    ws[f'E{row}'].number_format = currency_format
    ws[f'E{row}'].font = green_font
    
    row += 1
    ws[f'A{row}'] = "Closing Balance:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = cf_data['closing_balance']
    ws[f'B{row}'].number_format = currency_format
    ws[f'B{row}'].font = bold_font
    
    ws[f'D{row}'] = "Total Outflow:"
    ws[f'D{row}'].font = bold_font
    ws[f'E{row}'] = cf_data['total_outflow']
    ws[f'E{row}'].number_format = currency_format
    ws[f'E{row}'].font = red_font
    
    row += 1
    ws[f'D{row}'] = "Net Cash Flow:"
    ws[f'D{row}'].font = bold_font
    ws[f'E{row}'] = cf_data['net_cash_flow']
    ws[f'E{row}'].number_format = currency_format
    ws[f'E{row}'].font = bold_font
    
    row += 2
    
    # Transaction Headers
    headers = ['Date', 'Type', 'Description', 'Category', 'Reference', 'Inflow', 'Outflow', 'Account']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Opening Balance Row
    ws.cell(row=row, column=1, value=start_date.strftime('%d/%m/%Y'))
    ws.cell(row=row, column=2, value='B/F')
    ws.cell(row=row, column=3, value='Opening Balance')
    ws.cell(row=row, column=8, value=cf_data['opening_balance']).number_format = currency_format
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = gold_fill
    row += 1
    
    # Transactions
    for t in cf_data['transactions']:
        ws.cell(row=row, column=1, value=t['date'].strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=t['type'])
        ws.cell(row=row, column=3, value=t['description'])
        ws.cell(row=row, column=4, value=t['category'])
        ws.cell(row=row, column=5, value=t['reference'])
        
        if t['inflow']:
            cell = ws.cell(row=row, column=6, value=t['inflow'])
            cell.number_format = currency_format
            cell.font = green_font
        
        if t['outflow']:
            cell = ws.cell(row=row, column=7, value=t['outflow'])
            cell.number_format = currency_format
            cell.font = red_font
        
        ws.cell(row=row, column=8, value=t['account'])
        row += 1
    
    # Totals Row
    ws.cell(row=row, column=3, value='TOTALS').font = bold_font
    ws.cell(row=row, column=6, value=cf_data['total_inflow']).number_format = currency_format
    ws.cell(row=row, column=6).font = Font(bold=True, color="28A745")
    ws.cell(row=row, column=7, value=cf_data['total_outflow']).number_format = currency_format
    ws.cell(row=row, column=7).font = Font(bold=True, color="DC3545")
    row += 1
    
    # Closing Balance Row
    ws.cell(row=row, column=1, value=end_date.strftime('%d/%m/%Y'))
    ws.cell(row=row, column=2, value='C/F')
    ws.cell(row=row, column=3, value='Closing Balance')
    ws.cell(row=row, column=8, value=cf_data['closing_balance']).number_format = currency_format
    ws.cell(row=row, column=8).font = bold_font
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = gold_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    # Create response
    account_name = account.bank_name.replace(' ', '_') if account else 'All'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=CashFlow_{account_name}_{start_date}_{end_date}.xlsx'
    wb.save(response)
    return response



def get_customer_sites(request):
    """Get sites for a specific customer."""
    from .models import Site
    customer_id = request.GET.get('customer_id')
    if customer_id:
        sites = Site.objects.filter(customer_id=customer_id).values('id', 'name')
        return JsonResponse(list(sites), safe=False)
    return JsonResponse([], safe=False)


def get_customer_orders(request):
    """Get pending/partial sales orders for a specific customer."""
    from .models import SalesOrder
    customer_id = request.GET.get('customer_id')
    if customer_id:
        orders = SalesOrder.objects.filter(
            customer_id=customer_id,
            status__in=['PENDING', 'PARTIAL']
        ).values('id', 'date', 'pk')
        result = [{'id': o['id'], 'name': f"SO-{o['pk']:05d} ({o['date']})"} for o in orders]
        return JsonResponse(result, safe=False)
    return JsonResponse([], safe=False)


def get_order_items(request):
    """Get unfulfilled items for a specific sales order."""
    from .models import SalesOrderItem
    order_id = request.GET.get('order_id')
    if order_id:
        items = SalesOrderItem.objects.filter(order_id=order_id).select_related('block_type')
        result = []
        for item in items:
            remaining = item.quantity_requested - item.quantity_supplied
            if remaining > 0:
                result.append({
                    'id': item.id,
                    'name': f"{item.block_type.name} - {remaining} remaining @ ₦{item.agreed_price}"
                })
        return JsonResponse(result, safe=False)
    return JsonResponse([], safe=False)


def get_vendor_materials(request):
    """Get materials supplied by a specific vendor."""
    from .models import Vendor, Material
    vendor_id = request.GET.get('vendor_id')
    if vendor_id:
        vendor = Vendor.objects.filter(pk=vendor_id).first()
        if vendor and vendor.supply_type:
            material_map = {
                'CEMENT': ['CEMENT'],
                'SAND': ['SHARP_SAND', 'BLACK_SAND'],
                'TRANSPORT': [],
            }
            material_names = material_map.get(vendor.supply_type, [])
            materials = Material.objects.filter(name__in=material_names).values('id', 'name')
            result = [{'id': m['id'], 'name': m['name']} for m in materials]
            return JsonResponse(result, safe=False)
    materials = Material.objects.all().values('id', 'name')
    return JsonResponse(list(materials), safe=False)


