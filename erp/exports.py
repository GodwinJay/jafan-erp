import csv
from datetime import datetime, timedelta
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    SupplyLog, Expense, ProductionLog, Customer,
    Material, BlockType, Payment, ProcurementLog
)


class ReportExporter:
    """Handles all report exports for the ERP system."""

    def __init__(self, start_date=None, end_date=None):
        self.end_date = end_date or timezone.now().date()
        self.start_date = start_date or self.end_date.replace(day=1)

    # ==================== STYLING HELPERS ====================

    def _style_excel_header(self, ws, row=1):
        """Apply professional styling to Excel header row."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _auto_column_width(self, ws):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    def _add_summary_row(self, ws, label, value, row):
        """Add a summary row at the bottom."""
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).font = Font(bold=True)

    # ==================== SALES REPORT ====================

    def get_sales_data(self):
        """Fetch sales data for the date range."""
        return SupplyLog.objects.filter(
            date__gte=self.start_date,
            date__lte=self.end_date
        ).select_related('customer', 'site', 'block_type', 'truck', 'driver').order_by('-date')

    def export_sales_csv(self):
        """Export sales report as CSV."""
        response = HttpResponse(content_type='text/csv')
        response[
            'Content-Disposition'] = f'attachment; filename="sales_report_{self.start_date}_to_{self.end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Customer', 'Site', 'Block Type', 'Qty Loaded',
            'Breakages', 'Qty Delivered', 'Unit Price', 'Total Value',
            'Delivery Type', 'Truck', 'Driver'
        ])

        total_value = Decimal('0')
        total_delivered = 0

        for log in self.get_sales_data():
            writer.writerow([
                log.date,
                log.customer.name,
                log.site.name,
                log.block_type.name,
                log.quantity_loaded,
                log.breakages,
                log.quantity_delivered,
                log.unit_price,
                log.total_value,
                log.get_delivery_type_display(),
                log.truck.name if log.truck else 'N/A',
                log.driver.name if log.driver else 'N/A'
            ])
            total_value += log.total_value
            total_delivered += log.quantity_delivered

        writer.writerow([])
        writer.writerow(['TOTAL', '', '', '', '', '', total_delivered, '', total_value, '', '', ''])

        return response

    def export_sales_excel(self):
        """Export sales report as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Header
        headers = [
            'Date', 'Customer', 'Site', 'Block Type', 'Qty Loaded',
            'Breakages', 'Qty Delivered', 'Unit Price (₦)', 'Total Value (₦)',
            'Delivery Type', 'Truck', 'Driver'
        ]
        ws.append(headers)
        self._style_excel_header(ws)

        # Data
        total_value = Decimal('0')
        total_delivered = 0

        for log in self.get_sales_data():
            ws.append([
                log.date,
                log.customer.name,
                log.site.name,
                log.block_type.name,
                log.quantity_loaded,
                log.breakages,
                log.quantity_delivered,
                float(log.unit_price),
                float(log.total_value),
                log.get_delivery_type_display(),
                log.truck.name if log.truck else 'N/A',
                log.driver.name if log.driver else 'N/A'
            ])
            total_value += log.total_value
            total_delivered += log.quantity_delivered

        # Summary row
        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=7, value=total_delivered).font = Font(bold=True)
        ws.cell(row=summary_row, column=9, value=float(total_value)).font = Font(bold=True)

        self._auto_column_width(ws)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="sales_report_{self.start_date}_to_{self.end_date}.xlsx"'
        wb.save(response)
        return response

    # ==================== EXPENSES REPORT ====================

    def get_expenses_data(self):
        """Fetch expenses data for the date range."""
        return Expense.objects.filter(
            date__gte=self.start_date,
            date__lte=self.end_date
        ).select_related('category', 'payment_account', 'vendor', 'truck', 'employee').order_by('-date')

    def export_expenses_csv(self):
        """Export expenses report as CSV."""
        response = HttpResponse(content_type='text/csv')
        response[
            'Content-Disposition'] = f'attachment; filename="expenses_report_{self.start_date}_to_{self.end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Category', 'Description', 'Amount',
            'Payment Account', 'Vendor', 'Truck', 'Employee', 'Auto-Synced'
        ])

        total = Decimal('0')

        for exp in self.get_expenses_data():
            writer.writerow([
                exp.date,
                exp.category.name,
                exp.description,
                exp.amount,
                exp.payment_account.bank_name if exp.payment_account else 'N/A',
                exp.vendor.name if exp.vendor else 'N/A',
                exp.truck.name if exp.truck else 'N/A',
                exp.employee.name if exp.employee else 'N/A',
                'Yes' if exp.is_auto_synced else 'No'
            ])
            total += exp.amount

        writer.writerow([])
        writer.writerow(['TOTAL', '', '', total, '', '', '', '', ''])

        return response

    def export_expenses_excel(self):
        """Export expenses report as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses Report"

        headers = [
            'Date', 'Category', 'Description', 'Amount (₦)',
            'Payment Account', 'Vendor', 'Truck', 'Employee', 'Auto-Synced'
        ]
        ws.append(headers)
        self._style_excel_header(ws)

        total = Decimal('0')

        for exp in self.get_expenses_data():
            ws.append([
                exp.date,
                exp.category.name,
                exp.description,
                float(exp.amount),
                exp.payment_account.bank_name if exp.payment_account else 'N/A',
                exp.vendor.name if exp.vendor else 'N/A',
                exp.truck.name if exp.truck else 'N/A',
                exp.employee.name if exp.employee else 'N/A',
                'Yes' if exp.is_auto_synced else 'No'
            ])
            total += exp.amount

        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=4, value=float(total)).font = Font(bold=True)

        self._auto_column_width(ws)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="expenses_report_{self.start_date}_to_{self.end_date}.xlsx"'
        wb.save(response)
        return response

    # ==================== PRODUCTION REPORT ====================

    def get_production_data(self):
        """Fetch production data for the date range."""
        return ProductionLog.objects.filter(
            date__gte=self.start_date,
            date__lte=self.end_date
        ).select_related('team', 'machine', 'block_type').order_by('-date')

    def export_production_csv(self):
        """Export production report as CSV."""
        response = HttpResponse(content_type='text/csv')
        response[
            'Content-Disposition'] = f'attachment; filename="production_report_{self.start_date}_to_{self.end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Team', 'Machine', 'Block Type', 'Qty Produced',
            'Breakages', 'Cement Used', 'Sharp Sand Used', 'Black Sand Used', 'Labor Cost'
        ])

        total_produced = 0
        total_labor = Decimal('0')

        for log in self.get_production_data():
            writer.writerow([
                log.date,
                log.team.name,
                log.machine.name,
                log.block_type.name,
                log.quantity_produced,
                log.breakages,
                log.cement_used,
                log.sharp_sand_used,
                log.black_sand_used,
                log.labor_cost
            ])
            total_produced += log.quantity_produced
            total_labor += log.labor_cost

        writer.writerow([])
        writer.writerow(['TOTAL', '', '', '', total_produced, '', '', '', '', total_labor])

        return response

    def export_production_excel(self):
        """Export production report as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Production Report"

        headers = [
            'Date', 'Team', 'Machine', 'Block Type', 'Qty Produced',
            'Breakages', 'Cement Used', 'Sharp Sand Used', 'Black Sand Used', 'Labor Cost (₦)'
        ]
        ws.append(headers)
        self._style_excel_header(ws)

        total_produced = 0
        total_labor = Decimal('0')

        for log in self.get_production_data():
            ws.append([
                log.date,
                log.team.name,
                log.machine.name,
                log.block_type.name,
                log.quantity_produced,
                log.breakages,
                float(log.cement_used),
                float(log.sharp_sand_used),
                float(log.black_sand_used),
                float(log.labor_cost)
            ])
            total_produced += log.quantity_produced
            total_labor += log.labor_cost

        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=total_produced).font = Font(bold=True)
        ws.cell(row=summary_row, column=10, value=float(total_labor)).font = Font(bold=True)

        self._auto_column_width(ws)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="production_report_{self.start_date}_to_{self.end_date}.xlsx"'
        wb.save(response)
        return response

    # ==================== CUSTOMER LEDGER ====================

    def export_customer_ledger_excel(self, customer_id=None):
        """Export customer ledger - all customers or specific one."""
        wb = Workbook()

        customers = Customer.objects.filter(is_active=True)
        if customer_id:
            customers = customers.filter(id=customer_id)

        for customer in customers:
            ws = wb.create_sheet(title=customer.name[:30])

            # Customer Header
            ws.append([f"Customer Ledger: {customer.name}"])
            ws.append([f"Phone: {customer.phone}"])
            ws.append([f"Current Balance: ₦{customer.account_balance:,.2f}"])
            ws.append([f"Status: {customer.balance_status}"])
            ws.append([])

            # Transactions Header
            ws.append(['Date', 'Type', 'Description', 'Debit (₦)', 'Credit (₦)', 'Balance (₦)'])
            self._style_excel_header(ws, row=6)

            # Get all transactions
            transactions = []

            # Payments (Credits)
            for p in Payment.objects.filter(customer=customer, date__gte=self.start_date, date__lte=self.end_date):
                transactions.append({
                    'date': p.date,
                    'type': 'Payment',
                    'description': f'{p.get_method_display()} - {p.reference or "No Ref"}',
                    'debit': None,
                    'credit': p.amount
                })

            # Supplies (Debits)
            for s in SupplyLog.objects.filter(customer=customer, date__gte=self.start_date, date__lte=self.end_date):
                transactions.append({
                    'date': s.date,
                    'type': 'Supply',
                    'description': f'{s.quantity_delivered} x {s.block_type.name}',
                    'debit': s.total_value,
                    'credit': None
                })

            # Sort by date
            transactions.sort(key=lambda x: x['date'])

            # Write transactions
            running_balance = Decimal('0')
            for t in transactions:
                if t['credit']:
                    running_balance += t['credit']
                if t['debit']:
                    running_balance -= t['debit']

                ws.append([
                    t['date'],
                    t['type'],
                    t['description'],
                    float(t['debit']) if t['debit'] else '',
                    float(t['credit']) if t['credit'] else '',
                    float(running_balance)
                ])

            self._auto_column_width(ws)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="customer_ledger_{self.start_date}_to_{self.end_date}.xlsx"'
        wb.save(response)
        return response

    # ==================== INVENTORY REPORT ====================

    def export_inventory_excel(self):
        """Export current inventory status."""
        wb = Workbook()

        # Materials Sheet
        ws1 = wb.active
        ws1.title = "Raw Materials"
        ws1.append(['Material', 'Current Stock', 'Unit Price (₦)', 'Low Stock Threshold', 'Status'])
        self._style_excel_header(ws1)

        for mat in Material.objects.filter(is_active=True):
            status = "⚠️ LOW" if mat.is_low_stock else "✅ OK"
            ws1.append([
                mat.get_name_display(),
                float(mat.current_stock),
                float(mat.unit_price),
                float(mat.low_stock_threshold),
                status
            ])

        self._auto_column_width(ws1)

        # Block Types Sheet
        ws2 = wb.create_sheet(title="Block Inventory")
        ws2.append(['Block Type', 'Current Stock', 'Selling Price (₦)', 'Low Stock Threshold', 'Status'])
        self._style_excel_header(ws2)

        for block in BlockType.objects.filter(is_active=True):
            status = "⚠️ LOW" if block.is_low_stock else "✅ OK"
            ws2.append([
                block.name,
                block.current_stock,
                float(block.selling_price),
                block.low_stock_threshold,
                status
            ])

        self._auto_column_width(ws2)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{timezone.now().date()}.xlsx"'
        wb.save(response)
        return response